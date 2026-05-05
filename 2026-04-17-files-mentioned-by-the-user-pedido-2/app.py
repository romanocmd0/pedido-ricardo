from __future__ import annotations

import io
import hmac
import json
import os
import sqlite3
import unicodedata
from datetime import datetime
from datetime import date
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import qrcode
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image as ReportLabImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DATA_DIR = BASE_DIR / "data"
DATA_DIR = Path(os.getenv("DATA_DIR", str(DEFAULT_DATA_DIR))).resolve()
DATABASE_PATH = Path(os.getenv("DATABASE_PATH", str(DATA_DIR / "pedidos.db"))).resolve()
SEED_PATH = Path(os.getenv("SEED_PATH", str(DEFAULT_DATA_DIR / "seed_data.json"))).resolve()

MONTH_LABELS = {
    1: "JANEIRO",
    2: "FEVEREIRO",
    3: "MARCO",
    4: "ABRIL",
    5: "MAIO",
    6: "JUNHO",
    7: "JULHO",
    8: "AGOSTO",
    9: "SETEMBRO",
    10: "OUTUBRO",
    11: "NOVEMBRO",
    12: "DEZEMBRO",
}

MONTH_TITLES = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Marco",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro",
}

ALLOWED_SORT_COLUMNS = {
    "partner_name": "partner_name",
    "transferencia_qty": "transferencia_qty",
    "caminhao_transferencia_qty": "caminhao_transferencia_qty",
    "combo_transferencia_qty": "combo_transferencia_qty",
    "cautelar_qty": "cautelar_qty",
    "pesquisa_qty": "pesquisa_qty",
    "unit_transferencia": "unit_transferencia",
    "unit_caminhao_transferencia": "unit_caminhao_transferencia",
    "unit_combo_transferencia": "unit_combo_transferencia",
    "unit_cautelar": "unit_cautelar",
    "unit_pesquisa": "unit_pesquisa",
    "total_value": "total_value",
    "created_at": "created_at",
}

MIN_MONTH_KEY = "2026-04"
MAX_MONTH_KEY = "2050-12"
PIX_KEY = "64.683.079/0001-85"
PIX_KEY_VALUE = "64683079000185"
PIX_MERCHANT_NAME = "CERTIVE"
PIX_MERCHANT_CITY = "BRASIL"
CASH_SERVICE_CONFIG = {
    "TRANSFERENCIA": ("Transferencia", "transferencia_qty", "unit_transferencia"),
    "TRANSF. DE CAMINHAO": ("Transf. de Caminhao", "caminhao_transferencia_qty", "unit_caminhao_transferencia"),
    "TRANSFERENCIA DE CAMINHAO": ("Transf. de Caminhao", "caminhao_transferencia_qty", "unit_caminhao_transferencia"),
    "TRANSF. DO COMBO": ("Transf. do Combo", "combo_transferencia_qty", "unit_combo_transferencia"),
    "TRANSF. DE COMBO": ("Transf. do Combo", "combo_transferencia_qty", "unit_combo_transferencia"),
    "CAUTELAR": ("Cautelar", "cautelar_qty", "unit_cautelar"),
    "PESQUISA": ("Pesquisa", "pesquisa_qty", "unit_pesquisa"),
    "DIVERSOS": ("Diversos", None, None),
    "PAGAMENTO PARCEIROS": ("Pagamento Parceiros", None, None),
}

app = Flask(__name__)
app.config["JSON_AS_ASCII"] = False
app.secret_key = os.getenv("SECRET_KEY") or os.getenv("APP_SECRET_KEY") or os.urandom(32)

APP_PASSWORD = os.getenv("APP_PASSWORD", "")


def get_db() -> sqlite3.Connection:
    connection = sqlite3.connect(DATABASE_PATH)
    connection.row_factory = sqlite3.Row
    return connection


def is_authenticated() -> bool:
    return bool(session.get("authenticated"))


def wants_json_response() -> bool:
    return request.path.startswith("/api/")


@app.before_request
def require_login():
    public_endpoints = {"login", "healthcheck", "static"}
    if request.endpoint in public_endpoints:
        return None
    if is_authenticated():
        return None
    if wants_json_response():
        return jsonify({"error": "Autenticacao obrigatoria."}), 401
    return redirect(url_for("login", next=request.full_path if request.query_string else request.path))


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    normalized = unicodedata.normalize("NFKD", value.strip())
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    return ascii_only.upper()


def safe_filename_part(value: str) -> str:
    normalized = normalize_text(value).lower().replace(" ", "-")
    safe = "".join(character for character in normalized if character.isalnum() or character in {"-", "_"})
    return safe.strip("-") or "parceiro"


def normalize_client_name(value: str | None) -> str:
    if value is None:
        return ""
    cleaned_characters: list[str] = []
    for character in str(value):
        if unicodedata.category(character) in {"Cf", "Cc"}:
            continue
        if character in {"\u00A0", "\u2007", "\u202F"}:
            cleaned_characters.append(" ")
        else:
            cleaned_characters.append(character)
    return " ".join("".join(cleaned_characters).strip().split())


CLIENT_PRICE_FIELDS = {
    "Transferencia": "price_transferencia",
    "Transf. de Caminhao": "price_caminhao_transferencia",
    "Transf. do Combo": "price_combo_transferencia",
    "Cautelar": "price_cautelar",
    "Pesquisa": "price_pesquisa",
    "Diversos": "price_diversos",
}


def emv_field(field_id: str, value: str) -> str:
    return f"{field_id}{len(value):02d}{value}"


def crc16_ccitt(payload: str) -> str:
    crc = 0xFFFF
    for character in payload.encode("utf-8"):
        crc ^= character << 8
        for _ in range(8):
            if crc & 0x8000:
                crc = ((crc << 1) ^ 0x1021) & 0xFFFF
            else:
                crc = (crc << 1) & 0xFFFF
    return f"{crc:04X}"


def build_pix_payload() -> str:
    merchant_account = emv_field("00", "br.gov.bcb.pix") + emv_field("01", PIX_KEY_VALUE)
    payload_without_crc = "".join(
        [
            emv_field("00", "01"),
            emv_field("26", merchant_account),
            emv_field("52", "0000"),
            emv_field("53", "986"),
            emv_field("58", "BR"),
            emv_field("59", PIX_MERCHANT_NAME[:25]),
            emv_field("60", PIX_MERCHANT_CITY[:15]),
            emv_field("62", emv_field("05", "***")),
            "6304",
        ]
    )
    return f"{payload_without_crc}{crc16_ccitt(payload_without_crc)}"


def normalize_request_payment_status(value: str | None) -> str:
    normalized = normalize_text(value)
    if normalized == "PAGO":
        return "pago"
    return "em_aberto"


def request_payment_status_label(value: str | None) -> str:
    return "Pago" if normalize_request_payment_status(value) == "pago" else "Em aberto"


def parse_request_period(year_value: str | None, month_value: str | None) -> tuple[int | None, int | None]:
    year_number = None
    month_number = None
    if year_value:
        year_number = int(year_value)
        if year_number < 2000 or year_number > 2100:
            raise ValueError("Ano invalido.")
    if month_value:
        month_number = int(month_value)
        if month_number < 1 or month_number > 12:
            raise ValueError("Mes invalido.")
        if year_number is None:
            raise ValueError("Selecione o ano antes do mes.")
    return year_number, month_number


def month_key_from_parts(year_number: int, month_number: int) -> str:
    return f"{year_number:04d}-{month_number:02d}"


def title_from_month(month_number: int, year_number: int) -> str:
    return f"{MONTH_TITLES[month_number]} {year_number}"


def label_from_month(month_number: int) -> str:
    return MONTH_LABELS[month_number]


def parse_month_key(month_key: str) -> tuple[int, int]:
    try:
        year_raw, month_raw = month_key.split("-", 1)
        year_number = int(year_raw)
        month_number = int(month_raw)
    except (ValueError, AttributeError) as exc:
        raise ValueError("Mes invalido. Use o formato YYYY-MM.") from exc

    if month_number < 1 or month_number > 12:
        raise ValueError("Mes invalido. O valor do mes deve ficar entre 01 e 12.")
    return year_number, month_number


def parse_cash_date(cash_date: str) -> date:
    try:
        parsed = datetime.strptime(cash_date, "%Y-%m-%d").date()
    except (TypeError, ValueError) as exc:
        raise ValueError("Data invalida. Use o formato YYYY-MM-DD.") from exc
    return parsed


def format_display_date(cash_date: str) -> str:
    parsed = parse_cash_date(cash_date)
    return parsed.strftime("%d/%m/%Y")


def shift_month(year_number: int, month_number: int, offset: int) -> tuple[int, int]:
    absolute = (year_number * 12) + (month_number - 1) + offset
    next_year = absolute // 12
    next_month = (absolute % 12) + 1
    return next_year, next_month


def compare_month_key(month_key_a: str, month_key_b: str) -> int:
    year_a, month_a = parse_month_key(month_key_a)
    year_b, month_b = parse_month_key(month_key_b)
    if year_a == year_b and month_a == month_b:
        return 0
    return -1 if (year_a, month_a) < (year_b, month_b) else 1


def clamp_month_key(month_key: str) -> str:
    if compare_month_key(month_key, MIN_MONTH_KEY) < 0:
        return MIN_MONTH_KEY
    if compare_month_key(month_key, MAX_MONTH_KEY) > 0:
        return MAX_MONTH_KEY
    return month_key


def is_month_allowed(month_key: str) -> bool:
    return compare_month_key(month_key, MIN_MONTH_KEY) >= 0 and compare_month_key(month_key, MAX_MONTH_KEY) <= 0


def to_int(value: Any, field_name: str) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"O campo '{field_name}' deve ser inteiro.") from exc


def to_float(value: Any, field_name: str) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"O campo '{field_name}' deve ser numerico.") from exc


def calculate_total(record: dict[str, Any]) -> float:
    return round(
        (record["transferencia_qty"] * record["unit_transferencia"])
        + (record["caminhao_transferencia_qty"] * record["unit_caminhao_transferencia"])
        + (record["combo_transferencia_qty"] * record["unit_combo_transferencia"])
        + (record["cautelar_qty"] * record["unit_cautelar"])
        + (record["pesquisa_qty"] * record["unit_pesquisa"]),
        2,
    )


def validate_main_payload(payload: dict[str, Any]) -> dict[str, Any]:
    partner_name = (payload.get("partner_name") or "").strip()
    if not partner_name:
        raise ValueError("O nome do parceiro ou cliente e obrigatorio.")

    month_key = clamp_month_key((payload.get("month_key") or "").strip())
    if not month_key:
        raise ValueError("O mes ativo e obrigatorio.")
    if not is_month_allowed(month_key):
        raise ValueError("O mes precisa estar entre Abril de 2026 e Dezembro de 2050.")

    year_number, month_number = parse_month_key(month_key)
    record = {
        "month_key": month_key,
        "year_number": year_number,
        "month_number": month_number,
        "period_label": label_from_month(month_number),
        "partner_name": partner_name,
        "transferencia_qty": to_int(payload.get("transferencia_qty"), "transferencia_qty"),
        "caminhao_transferencia_qty": to_int(
            payload.get("caminhao_transferencia_qty"), "caminhao_transferencia_qty"
        ),
        "combo_transferencia_qty": to_int(payload.get("combo_transferencia_qty"), "combo_transferencia_qty"),
        "cautelar_qty": to_int(payload.get("cautelar_qty"), "cautelar_qty"),
        "pesquisa_qty": to_int(payload.get("pesquisa_qty"), "pesquisa_qty"),
        "unit_transferencia": to_float(payload.get("unit_transferencia"), "unit_transferencia"),
        "unit_caminhao_transferencia": to_float(
            payload.get("unit_caminhao_transferencia"), "unit_caminhao_transferencia"
        ),
        "unit_combo_transferencia": to_float(payload.get("unit_combo_transferencia"), "unit_combo_transferencia"),
        "unit_cautelar": to_float(payload.get("unit_cautelar"), "unit_cautelar"),
        "unit_pesquisa": to_float(payload.get("unit_pesquisa"), "unit_pesquisa"),
    }

    for field_name in (
        "transferencia_qty",
        "caminhao_transferencia_qty",
        "combo_transferencia_qty",
        "cautelar_qty",
        "pesquisa_qty",
        "unit_transferencia",
        "unit_caminhao_transferencia",
        "unit_combo_transferencia",
        "unit_cautelar",
        "unit_pesquisa",
    ):
        if record[field_name] < 0:
            raise ValueError(f"O campo '{field_name}' nao pode ser negativo.")

    record["total_value"] = calculate_total(record)
    return record


def normalize_payment_method(value: str) -> str:
    normalized = normalize_text(value)
    if "DINHEIRO" in normalized:
        return "dinheiro"
    if "DEBITO" in normalized or "DEBIT" in normalized:
        return "cartao_debito"
    if "CREDITO" in normalized or "CREDIT" in normalized:
        return "cartao_credito"
    if "PIX" in normalized:
        return "pix"
    return "outras"


def payment_label(payment_key: str) -> str:
    labels = {
        "dinheiro": "Dinheiro",
        "cartao_debito": "Cartao debito",
        "cartao_credito": "Cartao credito",
        "pix": "PIX",
        "outras": "Outras formas",
    }
    return labels.get(payment_key, "Outras formas")


def ensure_cash_day(connection: sqlite3.Connection, cash_date: str) -> None:
    parsed = parse_cash_date(cash_date)
    connection.execute(
        """
        INSERT OR IGNORE INTO cash_days (
            cash_date,
            year_number,
            month_number,
            day_number,
            month_title
        ) VALUES (?, ?, ?, ?, ?)
        """,
        (
            cash_date,
            parsed.year,
            parsed.month,
            parsed.day,
            title_from_month(parsed.month, parsed.year),
        ),
    )


def normalize_cash_service(value: str) -> str:
    normalized = normalize_text(value)
    if normalized not in CASH_SERVICE_CONFIG:
        raise ValueError("Selecione um servico valido.")
    return CASH_SERVICE_CONFIG[normalized][0]


def validate_cash_entry_payload(payload: dict[str, Any]) -> dict[str, Any]:
    customer_name = (payload.get("customer_name") or "").strip()
    service_name = (payload.get("service_name") or "").strip()
    payment_method = (payload.get("payment_method") or "").strip()
    flow_type = normalize_text(payload.get("flow_type") or "entrada").lower()
    if flow_type not in {"entrada", "saida", "deposito", "pagamento_parceiros"}:
        flow_type = "entrada"

    if not customer_name:
        raise ValueError("O nome do cliente e obrigatorio.")
    service_name = normalize_cash_service(service_name)
    if not payment_method:
        raise ValueError("A forma de pagamento e obrigatoria.")

    amount = to_float(payload.get("amount"), "amount")
    if amount < 0:
        raise ValueError("O valor nao pode ser negativo.")

    return {
        "customer_name": customer_name,
        "plate": (payload.get("plate") or "").strip().upper(),
        "service_name": service_name,
        "amount": round(amount, 2),
        "payment_method": payment_method,
        "payment_group": normalize_payment_method(payment_method),
        "flow_type": flow_type,
    }


def serialize_cash_entry(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "cash_date": row["cash_date"],
        "customer_name": row["customer_name"],
        "plate": row["plate"],
        "service_name": row["service_name"],
        "amount": row["amount"],
        "payment_method": row["payment_method"],
        "payment_group": row["payment_group"],
        "flow_type": row["flow_type"],
        "synced_to_monthly": bool(row["synced_to_monthly"]),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def cash_month_key(cash_date: str) -> str:
    parsed = parse_cash_date(cash_date)
    month_key = month_key_from_parts(parsed.year, parsed.month)
    if not is_month_allowed(month_key):
        raise ValueError("A data do caixa precisa estar entre Abril de 2026 e Dezembro de 2050.")
    return month_key


def monthly_record_payload_from_row(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "transferencia_qty": row["transferencia_qty"],
        "caminhao_transferencia_qty": row["caminhao_transferencia_qty"],
        "combo_transferencia_qty": row["combo_transferencia_qty"],
        "cautelar_qty": row["cautelar_qty"],
        "pesquisa_qty": row["pesquisa_qty"],
        "unit_transferencia": row["unit_transferencia"],
        "unit_caminhao_transferencia": row["unit_caminhao_transferencia"],
        "unit_combo_transferencia": row["unit_combo_transferencia"],
        "unit_cautelar": row["unit_cautelar"],
        "unit_pesquisa": row["unit_pesquisa"],
    }


def apply_monthly_cash_delta(
    connection: sqlite3.Connection,
    cash_date: str,
    customer_name: str,
    service_name: str,
    quantity_delta: int,
    amount_delta: float,
) -> None:
    if quantity_delta == 0 and amount_delta == 0:
        return

    service_config = CASH_SERVICE_CONFIG.get(normalize_text(service_name))
    if not service_config:
        return
    canonical_service, qty_field, unit_field = service_config
    if not qty_field or not unit_field:
        return

    month_key = cash_month_key(cash_date)
    year_number, month_number = parse_month_key(month_key)
    ensure_month(connection, year_number, month_number)
    row = connection.execute(
        """
        SELECT *
        FROM records
        WHERE month_key = ? AND partner_name = ?
        ORDER BY id
        LIMIT 1
        """,
        (month_key, customer_name),
    ).fetchone()

    if row is None:
        if quantity_delta <= 0:
            return
        record = {
            "month_key": month_key,
            "year_number": year_number,
            "month_number": month_number,
            "period_label": label_from_month(month_number),
            "partner_name": customer_name,
            "transferencia_qty": 0,
            "caminhao_transferencia_qty": 0,
            "combo_transferencia_qty": 0,
            "cautelar_qty": 0,
            "pesquisa_qty": 0,
            "unit_transferencia": 0,
            "unit_caminhao_transferencia": 0,
            "unit_combo_transferencia": 0,
            "unit_cautelar": 0,
            "unit_pesquisa": 0,
        }
        record[qty_field] = quantity_delta
        record[unit_field] = amount_delta / quantity_delta if quantity_delta else 0
        record["total_value"] = calculate_total(record)
        connection.execute(
            """
            INSERT INTO records (
                month_key,
                reference_date,
                period_label,
                period_sort,
                partner_name,
                transferencia_qty,
                caminhao_transferencia_qty,
                combo_transferencia_qty,
                cautelar_qty,
                pesquisa_qty,
                unit_transferencia,
                unit_caminhao_transferencia,
                unit_combo_transferencia,
                unit_cautelar,
                unit_pesquisa,
                total_value,
                updated_at
            ) VALUES (?, NULL, ?, 0, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            """,
            (
                record["month_key"],
                record["period_label"],
                record["partner_name"],
                record["transferencia_qty"],
                record["caminhao_transferencia_qty"],
                record["combo_transferencia_qty"],
                record["cautelar_qty"],
                record["pesquisa_qty"],
                record["unit_transferencia"],
                record["unit_caminhao_transferencia"],
                record["unit_combo_transferencia"],
                record["unit_cautelar"],
                record["unit_pesquisa"],
                record["total_value"],
            ),
        )
        return

    record = monthly_record_payload_from_row(row)
    current_qty = int(record[qty_field] or 0)
    current_total = float(current_qty * float(record[unit_field] or 0))
    next_qty = max(0, current_qty + quantity_delta)
    next_total = max(0.0, current_total + amount_delta)
    record[qty_field] = next_qty
    record[unit_field] = next_total / next_qty if next_qty else 0
    total_value = calculate_total(record)
    connection.execute(
        f"""
        UPDATE records
        SET
            {qty_field} = ?,
            {unit_field} = ?,
            total_value = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (record[qty_field], record[unit_field], total_value, row["id"]),
    )


def sync_cash_entry_to_monthly(connection: sqlite3.Connection, row: sqlite3.Row, direction: int) -> None:
    if row["flow_type"] != "entrada":
        return
    apply_monthly_cash_delta(
        connection,
        row["cash_date"],
        row["customer_name"],
        row["service_name"],
        direction,
        direction * float(row["amount"] or 0),
    )


def serialize_main_record(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "month_key": row["month_key"],
        "period_label": row["period_label"],
        "partner_name": row["partner_name"],
        "transferencia_qty": row["transferencia_qty"],
        "caminhao_transferencia_qty": row["caminhao_transferencia_qty"],
        "combo_transferencia_qty": row["combo_transferencia_qty"],
        "cautelar_qty": row["cautelar_qty"],
        "pesquisa_qty": row["pesquisa_qty"],
        "unit_transferencia": row["unit_transferencia"],
        "unit_caminhao_transferencia": row["unit_caminhao_transferencia"],
        "unit_combo_transferencia": row["unit_combo_transferencia"],
        "unit_cautelar": row["unit_cautelar"],
        "unit_pesquisa": row["unit_pesquisa"],
        "total_value": row["total_value"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def ensure_month(connection: sqlite3.Connection, year_number: int, month_number: int) -> str:
    month_key = month_key_from_parts(year_number, month_number)
    if not is_month_allowed(month_key):
        raise ValueError("O mes precisa estar entre Abril de 2026 e Dezembro de 2050.")
    connection.execute(
        """
        INSERT OR IGNORE INTO months (
            month_key,
            year_number,
            month_number,
            month_label,
            month_title
        ) VALUES (?, ?, ?, ?, ?)
        """,
        (
            month_key,
            year_number,
            month_number,
            label_from_month(month_number),
            title_from_month(month_number, year_number),
        ),
    )
    return month_key


def get_default_month_key() -> str:
    now = datetime.now()
    return clamp_month_key(month_key_from_parts(now.year, now.month))


def infer_seed_month_map(seed_records: list[dict[str, Any]]) -> dict[str, str]:
    ordered_labels: list[str] = []
    for item in seed_records:
        label = normalize_text(item.get("period_label"))
        if label and label not in ordered_labels:
            ordered_labels.append(label)
    if not ordered_labels:
        return {}

    current_year, current_month = parse_month_key(MIN_MONTH_KEY)
    month_map: dict[str, str] = {}
    for label in ordered_labels:
        month_key = month_key_from_parts(current_year, current_month)
        if is_month_allowed(month_key):
            month_map[label] = month_key
        current_year, current_month = shift_month(current_year, current_month, 1)
    return month_map


def ensure_allowed_months(connection: sqlite3.Connection) -> None:
    current_year, current_month = parse_month_key(MIN_MONTH_KEY)
    while compare_month_key(month_key_from_parts(current_year, current_month), MAX_MONTH_KEY) <= 0:
        ensure_month(connection, current_year, current_month)
        current_year, current_month = shift_month(current_year, current_month, 1)
    connection.execute("DELETE FROM months WHERE month_key < ? OR month_key > ?", (MIN_MONTH_KEY, MAX_MONTH_KEY))


def summarize_month(connection: sqlite3.Connection, month_key: str) -> dict[str, Any]:
    row = connection.execute(
        """
        SELECT
            COALESCE(SUM(total_value), 0) AS total_value,
            COALESCE(SUM(transferencia_qty), 0) AS transferencia_qty,
            COALESCE(SUM(caminhao_transferencia_qty), 0) AS caminhao_transferencia_qty,
            COALESCE(SUM(combo_transferencia_qty), 0) AS combo_transferencia_qty,
            COALESCE(SUM(cautelar_qty), 0) AS cautelar_qty,
            COALESCE(SUM(pesquisa_qty), 0) AS pesquisa_qty,
            COALESCE(SUM(transferencia_qty * unit_transferencia), 0) AS transferencia_total_value,
            COALESCE(SUM(caminhao_transferencia_qty * unit_caminhao_transferencia), 0) AS caminhao_transferencia_total_value,
            COALESCE(SUM(combo_transferencia_qty * unit_combo_transferencia), 0) AS combo_transferencia_total_value,
            COALESCE(SUM(cautelar_qty * unit_cautelar), 0) AS cautelar_total_value,
            COALESCE(SUM(pesquisa_qty * unit_pesquisa), 0) AS pesquisa_total_value,
            COUNT(*) AS record_count
        FROM records
        WHERE month_key = ?
        """,
        (month_key,),
    ).fetchone()

    total_operations = (
        row["transferencia_qty"]
        + row["caminhao_transferencia_qty"]
        + row["combo_transferencia_qty"]
        + row["cautelar_qty"]
        + row["pesquisa_qty"]
    )
    transferencia_group_qty = (
        row["transferencia_qty"] + row["caminhao_transferencia_qty"] + row["combo_transferencia_qty"]
    )
    transferencia_group_total_value = (
        row["transferencia_total_value"]
        + row["caminhao_transferencia_total_value"]
        + row["combo_transferencia_total_value"]
    )

    def percentage(value: int) -> float:
        if total_operations == 0:
            return 0.0
        return round((value / total_operations) * 100, 2)

    return {
        "total_value": row["total_value"],
        "record_count": row["record_count"],
        "transferencia_qty": row["transferencia_qty"],
        "caminhao_transferencia_qty": row["caminhao_transferencia_qty"],
        "combo_transferencia_qty": row["combo_transferencia_qty"],
        "cautelar_qty": row["cautelar_qty"],
        "pesquisa_qty": row["pesquisa_qty"],
        "transferencia_total_value": row["transferencia_total_value"],
        "caminhao_transferencia_total_value": row["caminhao_transferencia_total_value"],
        "combo_transferencia_total_value": row["combo_transferencia_total_value"],
        "cautelar_total_value": row["cautelar_total_value"],
        "pesquisa_total_value": row["pesquisa_total_value"],
        "transferencia_group_qty": transferencia_group_qty,
        "transferencia_group_total_value": transferencia_group_total_value,
        "total_operations": total_operations,
        "transferencia_pct": percentage(transferencia_group_qty),
        "caminhao_transferencia_pct": percentage(row["caminhao_transferencia_qty"]),
        "combo_transferencia_pct": percentage(row["combo_transferencia_qty"]),
        "cautelar_pct": percentage(row["cautelar_qty"]),
        "pesquisa_pct": percentage(row["pesquisa_qty"]),
    }


def list_months(connection: sqlite3.Connection) -> list[dict[str, Any]]:
    current_key = get_default_month_key()
    rows = connection.execute(
        """
        SELECT
            m.month_key,
            m.month_label,
            m.month_title,
            m.year_number,
            m.month_number,
            COUNT(r.id) AS record_count,
            COALESCE(SUM(r.total_value), 0) AS total_value
        FROM months m
        LEFT JOIN records r ON r.month_key = m.month_key
        GROUP BY
            m.month_key,
            m.month_label,
            m.month_title,
            m.year_number,
            m.month_number
        ORDER BY m.year_number, m.month_number
        """
    ).fetchall()

    return [
        {
            "month_key": row["month_key"],
            "month_label": row["month_label"],
            "month_title": row["month_title"],
            "year_number": row["year_number"],
            "month_number": row["month_number"],
            "record_count": row["record_count"],
            "total_value": row["total_value"],
            "is_current": row["month_key"] == current_key,
        }
        for row in rows
    ]


def get_month_title(connection: sqlite3.Connection, month_key: str) -> str:
    row = connection.execute("SELECT month_title FROM months WHERE month_key = ?", (month_key,)).fetchone()
    return row["month_title"] if row else month_key


def get_month_report(
    connection: sqlite3.Connection,
    month_key: str,
    search: str = "",
    sort_by: str = "partner_name",
    sort_order: str = "ASC",
) -> dict[str, Any]:
    conditions = ["month_key = ?"]
    params: list[Any] = [month_key]
    if search:
        conditions.append("partner_name LIKE ?")
        params.append(f"%{search}%")

    query = f"""
        SELECT *
        FROM records
        WHERE {' AND '.join(conditions)}
        ORDER BY {sort_by} {sort_order}, partner_name ASC, id DESC
    """
    rows = connection.execute(query, params).fetchall()
    return {
        "records": rows,
        "summary": summarize_month(connection, month_key),
        "month_title": get_month_title(connection, month_key),
        "month_key": month_key,
    }


def get_comparison_data(connection: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT
            m.month_key,
            m.month_title,
            COUNT(r.id) AS record_count,
            COALESCE(SUM(r.total_value), 0) AS total_value,
            COALESCE(SUM(r.transferencia_qty * r.unit_transferencia), 0) AS transferencia_total_value,
            COALESCE(SUM(r.caminhao_transferencia_qty * r.unit_caminhao_transferencia), 0) AS caminhao_transferencia_total_value,
            COALESCE(SUM(r.combo_transferencia_qty * r.unit_combo_transferencia), 0) AS combo_transferencia_total_value,
            COALESCE(SUM(r.cautelar_qty * r.unit_cautelar), 0) AS cautelar_total_value,
            COALESCE(SUM(r.pesquisa_qty * r.unit_pesquisa), 0) AS pesquisa_total_value
        FROM months m
        LEFT JOIN records r ON r.month_key = m.month_key
        GROUP BY m.month_key, m.month_title, m.year_number, m.month_number
        HAVING COUNT(r.id) > 0
        ORDER BY m.year_number, m.month_number
        """
    ).fetchall()

    return [
        {
            "month_key": row["month_key"],
            "month_title": row["month_title"],
            "record_count": row["record_count"],
            "total_value": row["total_value"],
            "transferencia_total_value": (
                row["transferencia_total_value"]
                + row["caminhao_transferencia_total_value"]
                + row["combo_transferencia_total_value"]
            ),
            "caminhao_transferencia_total_value": row["caminhao_transferencia_total_value"],
            "combo_transferencia_total_value": row["combo_transferencia_total_value"],
            "cautelar_total_value": row["cautelar_total_value"],
            "pesquisa_total_value": row["pesquisa_total_value"],
        }
        for row in rows
    ]


def get_client_price_payload(payload: dict[str, Any]) -> dict[str, float]:
    prices = {
        "price_transferencia": round(to_float(payload.get("price_transferencia"), "price_transferencia"), 2),
        "price_caminhao_transferencia": round(
            to_float(payload.get("price_caminhao_transferencia"), "price_caminhao_transferencia"), 2
        ),
        "price_combo_transferencia": round(
            to_float(payload.get("price_combo_transferencia"), "price_combo_transferencia"), 2
        ),
        "price_cautelar": round(to_float(payload.get("price_cautelar"), "price_cautelar"), 2),
        "price_pesquisa": round(to_float(payload.get("price_pesquisa"), "price_pesquisa"), 2),
        "price_diversos": round(to_float(payload.get("price_diversos"), "price_diversos"), 2),
    }
    if any(value < 0 for value in prices.values()):
        raise ValueError("Os valores do cadastro nao podem ser negativos.")
    return prices


def get_external_client_names(connection: sqlite3.Connection) -> list[str]:
    rows = connection.execute(
        """
        SELECT partner_name AS client_name
        FROM records
        WHERE TRIM(partner_name) <> ''
        UNION
        SELECT customer_name AS client_name
        FROM cash_entries
        WHERE TRIM(customer_name) <> ''
        ORDER BY client_name
        """
    ).fetchall()
    return [row["client_name"] for row in rows]


def get_all_clients(connection: sqlite3.Connection) -> list[str]:
    rows = connection.execute("SELECT client_name FROM clients WHERE TRIM(client_name) <> '' ORDER BY client_name").fetchall()
    return [row["client_name"] for row in rows]


def sync_clients_catalog(connection: sqlite3.Connection) -> None:
    current_rows = connection.execute("SELECT client_name FROM clients").fetchall()
    current_names = {normalize_text(row["client_name"]) for row in current_rows}
    for client_name in get_external_client_names(connection):
        if normalize_text(client_name) in current_names:
            continue
        upsert_client_from_system(connection, client_name)
        current_names.add(normalize_text(client_name))


def cleanup_clients_catalog(connection: sqlite3.Connection) -> dict[str, int]:
    rows = connection.execute(
        """
        SELECT
            id,
            client_name,
            price_transferencia,
            price_caminhao_transferencia,
            price_combo_transferencia,
            price_cautelar,
            price_pesquisa,
            price_diversos
        FROM clients
        ORDER BY id
        """
    ).fetchall()

    removed = 0
    updated = 0
    merged = 0
    normalized_map: dict[str, sqlite3.Row] = {}

    for row in rows:
        normalized_name = normalize_client_name(row["client_name"])
        if not normalized_name:
            connection.execute("DELETE FROM clients WHERE id = ?", (row["id"],))
            removed += 1
            continue

        normalized_key = normalize_text(normalized_name)
        survivor = normalized_map.get(normalized_key)
        if survivor is None:
            if normalized_name != row["client_name"]:
                connection.execute(
                    "UPDATE clients SET client_name = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    (normalized_name, row["id"]),
                )
                updated += 1
            normalized_map[normalized_key] = connection.execute("SELECT * FROM clients WHERE id = ?", (row["id"],)).fetchone()
            continue

        merged_values = {
            "price_transferencia": max(float(survivor["price_transferencia"] or 0), float(row["price_transferencia"] or 0)),
            "price_caminhao_transferencia": max(
                float(survivor["price_caminhao_transferencia"] or 0), float(row["price_caminhao_transferencia"] or 0)
            ),
            "price_combo_transferencia": max(
                float(survivor["price_combo_transferencia"] or 0), float(row["price_combo_transferencia"] or 0)
            ),
            "price_cautelar": max(float(survivor["price_cautelar"] or 0), float(row["price_cautelar"] or 0)),
            "price_pesquisa": max(float(survivor["price_pesquisa"] or 0), float(row["price_pesquisa"] or 0)),
            "price_diversos": max(float(survivor["price_diversos"] or 0), float(row["price_diversos"] or 0)),
        }
        connection.execute(
            """
            UPDATE clients
            SET
                price_transferencia = ?,
                price_caminhao_transferencia = ?,
                price_combo_transferencia = ?,
                price_cautelar = ?,
                price_pesquisa = ?,
                price_diversos = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                merged_values["price_transferencia"],
                merged_values["price_caminhao_transferencia"],
                merged_values["price_combo_transferencia"],
                merged_values["price_cautelar"],
                merged_values["price_pesquisa"],
                merged_values["price_diversos"],
                survivor["id"],
            ),
        )
        connection.execute("DELETE FROM clients WHERE id = ?", (row["id"],))
        normalized_map[normalized_key] = connection.execute("SELECT * FROM clients WHERE id = ?", (survivor["id"],)).fetchone()
        merged += 1

    return {"removed": removed, "updated": updated, "merged": merged}


def get_client_catalog(connection: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT
            id,
            client_name,
            price_transferencia,
            price_caminhao_transferencia,
            price_combo_transferencia,
            price_cautelar,
            price_pesquisa,
            price_diversos
        FROM clients
        WHERE TRIM(client_name) <> ''
        ORDER BY client_name
        """
    ).fetchall()
    return [
        {
            "id": row["id"],
            "client_name": row["client_name"],
            "price_transferencia": float(row["price_transferencia"] or 0),
            "price_caminhao_transferencia": float(row["price_caminhao_transferencia"] or 0),
            "price_combo_transferencia": float(row["price_combo_transferencia"] or 0),
            "price_cautelar": float(row["price_cautelar"] or 0),
            "price_pesquisa": float(row["price_pesquisa"] or 0),
            "price_diversos": float(row["price_diversos"] or 0),
        }
        for row in rows
    ]


def register_client(connection: sqlite3.Connection, client_name: str, payload: dict[str, Any] | None = None) -> dict[str, Any]:
    normalized_name = normalize_client_name(client_name)
    if not normalized_name:
        raise ValueError("O nome do cliente e obrigatorio.")

    rows = get_all_clients(connection)
    if any(normalize_text(item) == normalize_text(normalized_name) for item in rows):
        raise ValueError("Esse cliente ja esta cadastrado.")

    prices = get_client_price_payload(payload or {})
    connection.execute(
        """
        INSERT INTO clients (
            client_name,
            price_transferencia,
            price_caminhao_transferencia,
            price_combo_transferencia,
            price_cautelar,
            price_pesquisa,
            price_diversos,
            created_at,
            updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
        """,
        (
            normalized_name,
            prices["price_transferencia"],
            prices["price_caminhao_transferencia"],
            prices["price_combo_transferencia"],
            prices["price_cautelar"],
            prices["price_pesquisa"],
            prices["price_diversos"],
        ),
    )
    row = connection.execute("SELECT * FROM clients WHERE client_name = ?", (normalized_name,)).fetchone()
    return {
        "id": row["id"],
        "client_name": row["client_name"],
        "price_transferencia": float(row["price_transferencia"] or 0),
        "price_caminhao_transferencia": float(row["price_caminhao_transferencia"] or 0),
        "price_combo_transferencia": float(row["price_combo_transferencia"] or 0),
        "price_cautelar": float(row["price_cautelar"] or 0),
        "price_pesquisa": float(row["price_pesquisa"] or 0),
        "price_diversos": float(row["price_diversos"] or 0),
    }


def update_client(connection: sqlite3.Connection, client_id: int, client_name: str, payload: dict[str, Any] | None = None) -> dict[str, Any]:
    normalized_name = normalize_client_name(client_name)
    if not normalized_name:
        raise ValueError("O nome do cliente e obrigatorio.")
    existing = connection.execute("SELECT * FROM clients WHERE id = ?", (client_id,)).fetchone()
    if existing is None:
        raise ValueError("Cliente nao encontrado.")
    duplicate = connection.execute("SELECT id, client_name FROM clients WHERE id <> ?", (client_id,)).fetchall()
    if any(normalize_text(row["client_name"]) == normalize_text(normalized_name) for row in duplicate):
        raise ValueError("Esse cliente ja esta cadastrado.")
    prices = get_client_price_payload(payload or {})
    connection.execute(
        """
        UPDATE clients
        SET
            client_name = ?,
            price_transferencia = ?,
            price_caminhao_transferencia = ?,
            price_combo_transferencia = ?,
            price_cautelar = ?,
            price_pesquisa = ?,
            price_diversos = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (
            normalized_name,
            prices["price_transferencia"],
            prices["price_caminhao_transferencia"],
            prices["price_combo_transferencia"],
            prices["price_cautelar"],
            prices["price_pesquisa"],
            prices["price_diversos"],
            client_id,
        ),
    )
    row = connection.execute("SELECT * FROM clients WHERE id = ?", (client_id,)).fetchone()
    return {
        "id": row["id"],
        "client_name": row["client_name"],
        "price_transferencia": float(row["price_transferencia"] or 0),
        "price_caminhao_transferencia": float(row["price_caminhao_transferencia"] or 0),
        "price_combo_transferencia": float(row["price_combo_transferencia"] or 0),
        "price_cautelar": float(row["price_cautelar"] or 0),
        "price_pesquisa": float(row["price_pesquisa"] or 0),
        "price_diversos": float(row["price_diversos"] or 0),
    }


def upsert_client_from_system(connection: sqlite3.Connection, client_name: str) -> None:
    normalized_name = normalize_client_name(client_name)
    if not normalized_name:
        return
    existing = get_all_clients(connection)
    if any(normalize_text(item) == normalize_text(normalized_name) for item in existing):
        return
    connection.execute(
        """
        INSERT INTO clients (
            client_name,
            price_transferencia,
            price_caminhao_transferencia,
            price_combo_transferencia,
            price_cautelar,
            price_pesquisa,
            price_diversos,
            created_at,
            updated_at
        )
        VALUES (?, 0, 0, 0, 0, 0, 0, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
        """,
        (normalized_name,),
    )


def get_month_chart_data(connection: sqlite3.Connection, month_key: str) -> dict[str, Any]:
    summary = summarize_month(connection, month_key)
    month_title = get_month_title(connection, month_key)
    rows = connection.execute(
        """
        SELECT
            id,
            partner_name,
            (transferencia_qty * unit_transferencia) AS transferencia_total,
            (caminhao_transferencia_qty * unit_caminhao_transferencia) AS caminhao_transferencia_total,
            (combo_transferencia_qty * unit_combo_transferencia) AS combo_total,
            (cautelar_qty * unit_cautelar) AS cautelar_total,
            (pesquisa_qty * unit_pesquisa) AS pesquisa_total,
            total_value
        FROM records
        WHERE month_key = ?
        ORDER BY id
        """,
        (month_key,),
    ).fetchall()

    if not rows:
        return {
            "month_key": month_key,
            "month_title": month_title,
            "has_data": False,
            "pie": [],
            "quantity": [],
            "grouped_pie": [],
            "grouped_quantity": [],
        }

    pie_data = [
        {"label": "Transferencia", "value": summary["transferencia_total_value"]},
        {"label": "Transf. Caminhao", "value": summary["caminhao_transferencia_total_value"]},
        {"label": "Transf. de Combo", "value": summary["combo_transferencia_total_value"]},
        {"label": "Cautelar", "value": summary["cautelar_total_value"]},
        {"label": "Pesquisa", "value": summary["pesquisa_total_value"]},
    ]
    quantity_data = [
        {"label": "Transferencia", "value": summary["transferencia_qty"]},
        {"label": "Transf. Caminhao", "value": summary["caminhao_transferencia_qty"]},
        {"label": "Transf. de Combo", "value": summary["combo_transferencia_qty"]},
        {"label": "Cautelar", "value": summary["cautelar_qty"]},
        {"label": "Pesquisa", "value": summary["pesquisa_qty"]},
    ]
    grouped_pie_data = [
        {"label": "Transferencias", "value": summary["transferencia_group_total_value"]},
        {"label": "Cautelar", "value": summary["cautelar_total_value"]},
        {"label": "Pesquisa", "value": summary["pesquisa_total_value"]},
    ]
    grouped_quantity_data = [
        {"label": "Transferencias", "value": summary["transferencia_group_qty"]},
        {"label": "Cautelar", "value": summary["cautelar_qty"]},
        {"label": "Pesquisa", "value": summary["pesquisa_qty"]},
    ]

    return {
        "month_key": month_key,
        "month_title": month_title,
        "has_data": True,
        "pie": pie_data,
        "quantity": quantity_data,
        "grouped_pie": grouped_pie_data,
        "grouped_quantity": grouped_quantity_data,
    }


def get_client_history(connection: sqlite3.Connection, partner_name: str) -> dict[str, Any]:
    rows = connection.execute(
        """
        SELECT
            m.month_key,
            m.month_title,
            COALESCE(SUM(r.transferencia_qty + r.caminhao_transferencia_qty + r.combo_transferencia_qty + r.cautelar_qty + r.pesquisa_qty), 0) AS vistoria_count,
            COALESCE(SUM(r.total_value), 0) AS total_value
        FROM months m
        JOIN records r ON r.month_key = m.month_key
        WHERE r.partner_name = ?
        GROUP BY m.month_key, m.month_title, m.year_number, m.month_number
        ORDER BY m.year_number, m.month_number
        """,
        (partner_name,),
    ).fetchall()

    return {
        "partner_name": partner_name,
        "has_data": len(rows) > 0,
        "months": [
            {
                "month_key": row["month_key"],
                "month_title": row["month_title"],
                "vistoria_count": row["vistoria_count"],
                "total_value": row["total_value"],
            }
            for row in rows
        ],
    }


def build_month_metrics(connection: sqlite3.Connection, month_key: str) -> dict[str, Any]:
    summary = summarize_month(connection, month_key)
    return {
        "month_key": month_key,
        "month_title": get_month_title(connection, month_key),
        "total_value": summary["total_value"],
        "transferencia_total_value": summary["transferencia_group_total_value"],
        "caminhao_transferencia_total_value": summary["caminhao_transferencia_total_value"],
        "combo_transferencia_total_value": summary["combo_transferencia_total_value"],
        "cautelar_total_value": summary["cautelar_total_value"],
        "pesquisa_total_value": summary["pesquisa_total_value"],
        "transferencia_qty": summary["transferencia_group_qty"],
        "caminhao_transferencia_qty": summary["caminhao_transferencia_qty"],
        "combo_transferencia_qty": summary["combo_transferencia_qty"],
        "cautelar_qty": summary["cautelar_qty"],
        "pesquisa_qty": summary["pesquisa_qty"],
        "total_operations": summary["total_operations"],
    }


def calculate_delta(current_value: float, previous_value: float) -> dict[str, Any]:
    difference = round(current_value - previous_value, 2)
    if previous_value == 0:
        if current_value == 0:
            pct_change: float | None = 0.0
        else:
            pct_change = None
    else:
        pct_change = round((difference / previous_value) * 100, 2)
    return {"difference": difference, "pct_change": pct_change}


def compare_two_months(connection: sqlite3.Connection, month_key_a: str, month_key_b: str) -> dict[str, Any]:
    first = build_month_metrics(connection, month_key_a)
    second = build_month_metrics(connection, month_key_b)
    keys = [
        "total_value",
        "transferencia_total_value",
        "caminhao_transferencia_total_value",
        "combo_transferencia_total_value",
        "cautelar_total_value",
        "pesquisa_total_value",
        "total_operations",
        "transferencia_qty",
        "caminhao_transferencia_qty",
        "combo_transferencia_qty",
        "cautelar_qty",
        "pesquisa_qty",
    ]
    deltas = {key: calculate_delta(second[key], first[key]) for key in keys}
    return {"first": first, "second": second, "delta": deltas}


def get_client_ranking(connection: sqlite3.Connection, month_key: str | None) -> list[dict[str, Any]]:
    params: list[Any] = []
    condition = ""
    if month_key:
        condition = "WHERE month_key = ?"
        params.append(month_key)

    rows = connection.execute(
        f"""
        SELECT
            partner_name,
            COALESCE(SUM(total_value), 0) AS total_value,
            COALESCE(SUM(transferencia_qty + caminhao_transferencia_qty + combo_transferencia_qty + cautelar_qty + pesquisa_qty), 0) AS vistoria_count
        FROM records
        {condition}
        GROUP BY partner_name
        HAVING TRIM(partner_name) <> ''
        ORDER BY total_value DESC, vistoria_count DESC, partner_name ASC
        """,
        params,
    ).fetchall()

    return [
        {
            "partner_name": row["partner_name"],
            "total_value": row["total_value"],
            "vistoria_count": row["vistoria_count"],
        }
        for row in rows
    ]


def get_client_comparison_payload(connection: sqlite3.Connection, month_key: str | None) -> dict[str, Any]:
    ranking = get_client_ranking(connection, month_key)
    top_clients = ranking[:8]

    if month_key:
        months_filter = "WHERE m.month_key = ?"
        params: list[Any] = [month_key]
    else:
        months_filter = """
        WHERE EXISTS (
            SELECT 1
            FROM records r
            WHERE r.month_key = m.month_key
        )
        """
        params = []

    months = connection.execute(
        f"""
        SELECT m.month_key, m.month_title
        FROM months m
        {months_filter}
        ORDER BY m.year_number, m.month_number
        """,
        params,
    ).fetchall()

    top_names = [item["partner_name"] for item in top_clients[:5]]
    evolution: list[dict[str, Any]] = []
    for month in months:
        row = {"month_key": month["month_key"], "month_title": month["month_title"]}
        for name in top_names:
            total = connection.execute(
                """
                SELECT COALESCE(SUM(total_value), 0) AS total_value
                FROM records
                WHERE month_key = ? AND partner_name = ?
                """,
                (month["month_key"], name),
            ).fetchone()["total_value"]
            row[name] = total
        evolution.append(row)

    return {
        "scope": month_key or "all",
        "ranking": ranking,
        "top_names": top_names,
        "evolution": evolution,
    }


def summarize_cash_day(connection: sqlite3.Connection, cash_date: str) -> dict[str, Any]:
    rows = connection.execute(
        """
        SELECT payment_group, flow_type, COALESCE(SUM(amount), 0) AS total_value, COUNT(*) AS entry_count
        FROM cash_entries
        WHERE cash_date = ?
        GROUP BY payment_group, flow_type
        """,
        (cash_date,),
    ).fetchall()

    payment_totals = {
        "dinheiro": 0.0,
        "cartao_debito": 0.0,
        "cartao_credito": 0.0,
        "pix": 0.0,
        "outras": 0.0,
    }
    total_in = 0.0
    total_out = 0.0
    total_deposit = 0.0
    total_partner_payment = 0.0
    entry_count = 0
    for row in rows:
        value = float(row["total_value"] or 0)
        entry_count += int(row["entry_count"] or 0)
        if row["flow_type"] == "saida":
            total_out += value
        elif row["flow_type"] == "deposito":
            total_deposit += value
            total_out += value
        elif row["flow_type"] == "pagamento_parceiros":
            total_partner_payment += value
        else:
            total_in += value
            payment_totals[row["payment_group"] if row["payment_group"] in payment_totals else "outras"] += value

    vault_row = connection.execute(
        """
        SELECT
            COALESCE(SUM(CASE
                WHEN flow_type = 'entrada' AND payment_group = 'dinheiro' THEN amount
                WHEN flow_type IN ('saida', 'deposito') THEN -amount
                ELSE 0
            END), 0) AS vault_balance
        FROM cash_entries
        WHERE cash_date <= ?
        """,
        (cash_date,),
    ).fetchone()

    return {
        "cash_date": cash_date,
        "display_date": format_display_date(cash_date),
        "payment_totals": payment_totals,
        "total_in": round(total_in, 2),
        "total_out": round(total_out, 2),
        "total_deposit": round(total_deposit, 2),
        "total_partner_payment": round(total_partner_payment, 2),
        "result": round(total_in - total_out, 2),
        "vault_balance": round(float(vault_row["vault_balance"] or 0), 2),
        "entry_count": entry_count,
    }


def get_cash_day_payload(connection: sqlite3.Connection, cash_date: str) -> dict[str, Any]:
    ensure_cash_day(connection, cash_date)
    day = connection.execute("SELECT * FROM cash_days WHERE cash_date = ?", (cash_date,)).fetchone()
    entries = connection.execute(
        """
        SELECT *
        FROM cash_entries
        WHERE cash_date = ?
        ORDER BY id
        """,
        (cash_date,),
    ).fetchall()
    return {
        "day": {
            "cash_date": day["cash_date"],
            "display_date": format_display_date(day["cash_date"]),
            "year_number": day["year_number"],
            "month_number": day["month_number"],
            "month_title": day["month_title"],
            "finalized": bool(day["finalized"]),
            "finalized_at": day["finalized_at"],
        },
        "entries": [serialize_cash_entry(row) for row in entries],
        "summary": summarize_cash_day(connection, cash_date),
    }


def get_cash_month_payload(connection: sqlite3.Connection, month_key: str) -> dict[str, Any]:
    year_number, month_number = parse_month_key(month_key)
    days = connection.execute(
        """
        SELECT
            d.cash_date,
            d.finalized,
            COUNT(e.id) AS entry_count
        FROM cash_days d
        LEFT JOIN cash_entries e ON e.cash_date = d.cash_date
        WHERE d.year_number = ? AND d.month_number = ?
        GROUP BY d.cash_date, d.finalized
        ORDER BY d.cash_date
        """,
        (year_number, month_number),
    ).fetchall()

    day_rows = []
    totals = {
        "total_in": 0.0,
        "total_out": 0.0,
        "total_deposit": 0.0,
        "total_partner_payment": 0.0,
        "result": 0.0,
        "entry_count": 0,
    }
    for day in days:
        summary = summarize_cash_day(connection, day["cash_date"])
        day_payload = {
            "cash_date": day["cash_date"],
            "display_date": format_display_date(day["cash_date"]),
            "finalized": bool(day["finalized"]),
            "entry_count": int(day["entry_count"] or 0),
            "summary": summary,
        }
        day_rows.append(day_payload)
        totals["total_in"] += summary["total_in"]
        totals["total_out"] += summary["total_out"]
        totals["total_deposit"] += summary["total_deposit"]
        totals["total_partner_payment"] += summary["total_partner_payment"]
        totals["result"] += summary["result"]
        totals["entry_count"] += summary["entry_count"]

    return {
        "month_key": month_key,
        "month_title": f"{MONTH_TITLES[month_number]} de {year_number}",
        "days": day_rows,
        "totals": {key: round(value, 2) for key, value in totals.items()},
    }


def get_partner_request_rows(
    connection: sqlite3.Connection,
    partner_name: str | None = None,
    year_number: int | None = None,
    month_number: int | None = None,
    only_open: bool = False,
) -> list[sqlite3.Row]:
    params: list[Any] = []
    partner_filter = ""
    if partner_name:
        partner_filter = "AND customer_name = ?"
        params.append(partner_name)
    period_filter = ""
    if year_number is not None:
        period_filter += " AND SUBSTR(cash_date, 1, 4) = ?"
        params.append(f"{year_number:04d}")
    if month_number is not None:
        period_filter += " AND SUBSTR(cash_date, 6, 2) = ?"
        params.append(f"{month_number:02d}")
    status_filter = ""
    if only_open:
        status_filter = "AND COALESCE(request_payment_status, 'em_aberto') = 'em_aberto'"

    return connection.execute(
        f"""
        SELECT
            id,
            cash_date,
            customer_name,
            plate,
            service_name,
            amount,
            payment_method,
            flow_type,
            COALESCE(request_payment_status, 'em_aberto') AS request_payment_status,
            created_at,
            updated_at
        FROM cash_entries
        WHERE UPPER(TRIM(payment_method)) = 'REC'
        {partner_filter}
        {period_filter}
        {status_filter}
        ORDER BY customer_name COLLATE NOCASE, cash_date, id
        """,
        params,
    ).fetchall()


def get_partner_requests_payload(connection: sqlite3.Connection) -> dict[str, Any]:
    rows = get_partner_request_rows(connection)
    partners: dict[str, dict[str, Any]] = {}
    for row in rows:
        partner = row["customer_name"]
        parsed = parse_cash_date(row["cash_date"])
        item = partners.setdefault(
            partner,
            {
                "partner_name": partner,
                "entry_count": 0,
                "total_value": 0.0,
                "last_date": row["cash_date"],
                "years": {},
            },
        )
        item["entry_count"] += 1
        item["total_value"] += float(row["amount"] or 0)
        item["last_date"] = max(item["last_date"], row["cash_date"])
        year_node = item["years"].setdefault(
            parsed.year,
            {
                "year_number": parsed.year,
                "entry_count": 0,
                "total_value": 0.0,
                "months": {},
            },
        )
        year_node["entry_count"] += 1
        year_node["total_value"] += float(row["amount"] or 0)
        month_node = year_node["months"].setdefault(
            parsed.month,
            {
                "month_number": parsed.month,
                "month_title": MONTH_TITLES[parsed.month],
                "entry_count": 0,
                "total_value": 0.0,
            },
        )
        month_node["entry_count"] += 1
        month_node["total_value"] += float(row["amount"] or 0)

    return {
        "pix_key": PIX_KEY,
        "pix_copy_paste": build_pix_payload(),
        "partners": [
            {
                "partner_name": partner["partner_name"],
                "entry_count": partner["entry_count"],
                "last_date": partner["last_date"],
                "total_value": round(partner["total_value"], 2),
                "last_display_date": format_display_date(partner["last_date"]),
                "years": [
                    {
                        "year_number": year["year_number"],
                        "entry_count": year["entry_count"],
                        "total_value": round(year["total_value"], 2),
                        "months": [
                            {
                                "month_number": month["month_number"],
                                "month_title": month["month_title"],
                                "entry_count": month["entry_count"],
                                "total_value": round(month["total_value"], 2),
                            }
                            for month in sorted(year["months"].values(), key=lambda item: item["month_number"], reverse=True)
                        ],
                    }
                    for year in sorted(partner["years"].values(), key=lambda item: item["year_number"], reverse=True)
                ],
            }
            for partner in sorted(partners.values(), key=lambda item: item["partner_name"].lower())
        ],
    }


def get_partner_request_detail(
    connection: sqlite3.Connection,
    partner_name: str,
    year_number: int | None = None,
    month_number: int | None = None,
    only_open: bool = False,
) -> dict[str, Any]:
    rows = get_partner_request_rows(
        connection,
        partner_name,
        year_number=year_number,
        month_number=month_number,
        only_open=only_open,
    )
    entries = [
        {
            "id": row["id"],
            "cash_date": row["cash_date"],
            "display_date": format_display_date(row["cash_date"]),
            "plate": row["plate"],
            "service_name": row["service_name"],
            "amount": round(float(row["amount"] or 0), 2),
            "payment_method": row["payment_method"],
            "flow_type": row["flow_type"],
            "request_payment_status": normalize_request_payment_status(row["request_payment_status"]),
            "request_payment_status_label": request_payment_status_label(row["request_payment_status"]),
        }
        for row in rows
    ]
    return {
        "pix_key": PIX_KEY,
        "pix_copy_paste": build_pix_payload(),
        "partner_name": partner_name,
        "title": partner_name.upper(),
        "selected_year": year_number,
        "selected_month": month_number,
        "selected_month_title": MONTH_TITLES[month_number] if month_number else "",
        "selected_period_label": f"{MONTH_TITLES[month_number]} de {year_number}" if year_number and month_number else "",
        "entries": entries,
        "entry_count": len(entries),
        "total_value": round(sum(entry["amount"] for entry in entries), 2),
    }


def get_cash_tree(connection: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT
            d.cash_date,
            d.year_number,
            d.month_number,
            d.month_title,
            d.finalized,
            COUNT(e.id) AS entry_count,
            COALESCE(SUM(CASE
                WHEN e.flow_type = 'entrada' THEN e.amount
                WHEN e.flow_type IN ('saida', 'deposito') THEN -e.amount
                ELSE 0
            END), 0) AS result_value
        FROM cash_days d
        LEFT JOIN cash_entries e ON e.cash_date = d.cash_date
        GROUP BY d.cash_date, d.year_number, d.month_number, d.month_title, d.finalized
        ORDER BY d.year_number DESC, d.month_number DESC, d.day_number DESC
        """
    ).fetchall()

    tree: dict[int, dict[str, Any]] = {}
    for row in rows:
        year = row["year_number"]
        month = row["month_number"]
        year_node = tree.setdefault(year, {"year": year, "months": {}})
        month_node = year_node["months"].setdefault(
            month,
            {
                "month_number": month,
                "month_title": row["month_title"],
                "days": [],
            },
        )
        month_node["days"].append(
            {
                "cash_date": row["cash_date"],
                "display_date": format_display_date(row["cash_date"]),
                "entry_count": row["entry_count"],
                "result_value": row["result_value"],
                "finalized": bool(row["finalized"]),
            }
        )

    return [
        {
            "year": year_node["year"],
            "months": list(year_node["months"].values()),
        }
        for year_node in tree.values()
    ]


def build_excel_report(report: dict[str, Any]) -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Relatorio"

    title_fill = PatternFill("solid", fgColor="0E2A47")
    title_font = Font(color="FFFFFF", bold=True, size=13)
    header_fill = PatternFill("solid", fgColor="D4AF37")
    header_font = Font(bold=True, color="0E2A47")

    sheet.merge_cells("A1:L1")
    sheet["A1"] = f"Relatorio Mensal - {report['month_title']}"
    sheet["A1"].fill = title_fill
    sheet["A1"].font = title_font
    sheet["A1"].alignment = Alignment(horizontal="center")

    summary = report["summary"]
    summary_rows = [
        ("Valor total", summary["total_value"]),
        ("Total de Transferencias", summary["transferencia_group_total_value"]),
        ("Total Transf. de Combo", summary["combo_transferencia_total_value"]),
        ("Total Cautelar", summary["cautelar_total_value"]),
        ("Total Pesquisa", summary["pesquisa_total_value"]),
        ("Registros no mes", summary["record_count"]),
        ("Total de operacoes", summary["total_operations"]),
        ("Qtd. Transferencias", summary["transferencia_group_qty"]),
        ("Qtd. Transf. Caminhao", summary["caminhao_transferencia_qty"]),
        ("Qtd. Transf. de Combo", summary["combo_transferencia_qty"]),
        ("Qtd. Cautelares", summary["cautelar_qty"]),
        ("Qtd. Pesquisas", summary["pesquisa_qty"]),
        ("Percentual Transferencias", f"{summary['transferencia_pct']}%"),
        ("Percentual Transf. Caminhao", f"{summary['caminhao_transferencia_pct']}%"),
        ("Percentual Transf. de Combo", f"{summary['combo_transferencia_pct']}%"),
        ("Percentual Cautelares", f"{summary['cautelar_pct']}%"),
        ("Percentual Pesquisas", f"{summary['pesquisa_pct']}%"),
    ]

    current_row = 3
    for label, value in summary_rows:
        sheet[f"A{current_row}"] = label
        sheet[f"B{current_row}"] = value
        current_row += 1

    current_row += 1
    headers = [
        "Parceiro",
        "Transfer.",
        "Transf. Caminhao",
        "Transf. de Combo",
        "Cautelar",
        "Pesquisa",
        "Vlr. Transfer.",
        "Vlr. Caminhao",
        "Vlr. Combo",
        "Vlr. Cautelar",
        "Vlr. Pesquisa",
        "Total",
    ]
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=current_row, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row in report["records"]:
        current_row += 1
        values = [
            row["partner_name"],
            row["transferencia_qty"],
            row["caminhao_transferencia_qty"],
            row["combo_transferencia_qty"],
            row["cautelar_qty"],
            row["pesquisa_qty"],
            row["unit_transferencia"],
            row["unit_caminhao_transferencia"],
            row["unit_combo_transferencia"],
            row["unit_cautelar"],
            row["unit_pesquisa"],
            row["total_value"],
        ]
        for col_index, value in enumerate(values, start=1):
            sheet.cell(row=current_row, column=col_index, value=value)

    for column_letter, width in {
        "A": 28,
        "B": 12,
        "C": 18,
        "D": 18,
        "E": 12,
        "F": 12,
        "G": 15,
        "H": 15,
        "I": 15,
        "J": 15,
        "K": 15,
        "L": 15,
    }.items():
        sheet.column_dimensions[column_letter].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_pdf_report(report: dict[str, Any]) -> io.BytesIO:
    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    story = [Paragraph(f"Relatorio Mensal - {report['month_title']}", styles["Title"]), Spacer(1, 8)]

    summary = report["summary"]
    summary_data = [
        ["Metrica", "Valor"],
        ["Valor total", f"R$ {summary['total_value']:.2f}"],
        ["Total de Transferencias", f"R$ {summary['transferencia_group_total_value']:.2f}"],
        ["Total Transf. de Combo", f"R$ {summary['combo_transferencia_total_value']:.2f}"],
        ["Total Cautelar", f"R$ {summary['cautelar_total_value']:.2f}"],
        ["Total Pesquisa", f"R$ {summary['pesquisa_total_value']:.2f}"],
        ["Registros no mes", str(summary["record_count"])],
        ["Total de operacoes", str(summary["total_operations"])],
        ["Qtd. Transferencias", str(summary["transferencia_group_qty"])],
        ["Qtd. Transf. Caminhao", str(summary["caminhao_transferencia_qty"])],
        ["Qtd. Transf. de Combo", str(summary["combo_transferencia_qty"])],
        ["Qtd. Cautelares", str(summary["cautelar_qty"])],
        ["Qtd. Pesquisas", str(summary["pesquisa_qty"])],
        ["Percentual Transferencias", f"{summary['transferencia_pct']}%"],
        ["Percentual Transf. Caminhao", f"{summary['caminhao_transferencia_pct']}%"],
        ["Percentual Transf. de Combo", f"{summary['combo_transferencia_pct']}%"],
        ["Percentual Cautelares", f"{summary['cautelar_pct']}%"],
        ["Percentual Pesquisas", f"{summary['pesquisa_pct']}%"],
    ]
    summary_table = Table(summary_data, colWidths=[70 * mm, 45 * mm])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E2A47")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D4AF37")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8F4E6")),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 10)])

    table_data = [[
        "Parceiro",
        "Transfer.",
        "Caminhao",
        "Combo",
        "Cautelar",
        "Pesquisa",
        "Vlr. Transfer.",
        "Vlr. Caminhao",
        "Vlr. Combo",
        "Vlr. Cautelar",
        "Vlr. Pesquisa",
        "Total",
    ]]
    for row in report["records"]:
        table_data.append(
            [
                row["partner_name"],
                row["transferencia_qty"],
                row["caminhao_transferencia_qty"],
                row["combo_transferencia_qty"],
                row["cautelar_qty"],
                row["pesquisa_qty"],
                f"R$ {row['unit_transferencia']:.2f}",
                f"R$ {row['unit_caminhao_transferencia']:.2f}",
                f"R$ {row['unit_combo_transferencia']:.2f}",
                f"R$ {row['unit_cautelar']:.2f}",
                f"R$ {row['unit_pesquisa']:.2f}",
                f"R$ {row['total_value']:.2f}",
            ]
        )

    table = Table(
        table_data,
        colWidths=[
            39 * mm,
            14 * mm,
            16 * mm,
            16 * mm,
            14 * mm,
            14 * mm,
            19 * mm,
            19 * mm,
            19 * mm,
            19 * mm,
            19 * mm,
            19 * mm,
        ],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E2A47")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D4AF37")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FFFDF7")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(table)
    document.build(story)
    buffer.seek(0)
    return buffer


def build_cash_pdf_report(payload: dict[str, Any]) -> io.BytesIO:
    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    day = payload["day"]
    summary = payload["summary"]
    story = [Paragraph(f"Fluxo de Caixa - {day['display_date']}", styles["Title"]), Spacer(1, 8)]

    summary_data = [
        ["Metrica", "Valor"],
        ["Dinheiro", f"R$ {summary['payment_totals']['dinheiro']:.2f}"],
        ["Cartao debito", f"R$ {summary['payment_totals']['cartao_debito']:.2f}"],
        ["Cartao credito", f"R$ {summary['payment_totals']['cartao_credito']:.2f}"],
        ["PIX", f"R$ {summary['payment_totals']['pix']:.2f}"],
        ["Outras formas", f"R$ {summary['payment_totals']['outras']:.2f}"],
        ["Total de entradas", f"R$ {summary['total_in']:.2f}"],
        ["Total de saidas", f"R$ {summary['total_out']:.2f}"],
        ["Depositos", f"R$ {summary['total_deposit']:.2f}"],
        ["Pagamento Parceiros", f"R$ {summary['total_partner_payment']:.2f}"],
        ["Resultado do dia", f"R$ {summary['result']:.2f}"],
        ["Cofre atual", f"R$ {summary['vault_balance']:.2f}"],
        ["Status", "Finalizado" if day["finalized"] else "Aberto"],
    ]
    summary_table = Table(summary_data, colWidths=[70 * mm, 45 * mm])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E2A47")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D4AF37")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8F4E6")),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 10)])

    table_data = [["Cliente", "Placa", "Servico", "Valor", "Pagamento", "Tipo"]]
    for entry in payload["entries"]:
        flow_label = {
            "entrada": "Entrada",
            "saida": "Saida",
            "deposito": "Deposito",
            "pagamento_parceiros": "Pagamento Parceiros",
        }.get(entry["flow_type"], "Entrada")
        table_data.append(
            [
                entry["customer_name"],
                entry["plate"],
                entry["service_name"],
                f"R$ {entry['amount']:.2f}",
                entry["payment_method"],
                flow_label,
            ]
        )

    table = Table(table_data, colWidths=[54 * mm, 28 * mm, 38 * mm, 26 * mm, 42 * mm, 28 * mm], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E2A47")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D4AF37")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FFFDF7")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(table)
    document.build(story)
    buffer.seek(0)
    return buffer


def build_cash_month_pdf_report(payload: dict[str, Any]) -> io.BytesIO:
    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    totals = payload["totals"]
    story = [Paragraph(f"Fluxo de Caixa Mensal - {payload['month_title']}", styles["Title"]), Spacer(1, 8)]

    summary_data = [
        ["Resumo do mes", "Valor"],
        ["Total de entradas", f"R$ {totals['total_in']:.2f}"],
        ["Total de saidas", f"R$ {totals['total_out']:.2f}"],
        ["Depositos", f"R$ {totals['total_deposit']:.2f}"],
        ["Pagamento Parceiros", f"R$ {totals['total_partner_payment']:.2f}"],
        ["Resultado acumulado", f"R$ {totals['result']:.2f}"],
        ["Total de lancamentos", str(int(totals["entry_count"]))],
    ]
    summary_table = Table(summary_data, colWidths=[70 * mm, 45 * mm])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E2A47")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D4AF37")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8F4E6")),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 10)])

    table_data = [["Dia", "Lancamentos", "Entradas", "Saidas", "Depositos", "Pgto. Parceiros", "Resultado", "Status"]]
    if payload["days"]:
        for day in payload["days"]:
            summary = day["summary"]
            table_data.append(
                [
                    day["display_date"],
                    str(day["entry_count"]),
                    f"R$ {summary['total_in']:.2f}",
                    f"R$ {summary['total_out']:.2f}",
                    f"R$ {summary['total_deposit']:.2f}",
                    f"R$ {summary['total_partner_payment']:.2f}",
                    f"R$ {summary['result']:.2f}",
                    "Finalizado" if day["finalized"] else "Aberto",
                ]
            )
    else:
        table_data.append(["Sem caixas no mes", "0", "R$ 0.00", "R$ 0.00", "R$ 0.00", "R$ 0.00", "R$ 0.00", "-"])

    table = Table(table_data, colWidths=[27 * mm, 23 * mm, 28 * mm, 28 * mm, 28 * mm, 32 * mm, 30 * mm, 24 * mm], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E2A47")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D4AF37")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FFFDF7")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(table)
    document.build(story)
    buffer.seek(0)
    return buffer


def build_pix_qr_png() -> io.BytesIO:
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=4,
    )
    qr.add_data(build_pix_payload())
    qr.make(fit=True)
    image = qr.make_image(fill_color="black", back_color="white")
    buffer = io.BytesIO()
    image.save(buffer, format="PNG")
    buffer.seek(0)
    return buffer


def build_pix_qr_flowable(size: int = 72) -> ReportLabImage:
    buffer = build_pix_qr_png()
    return ReportLabImage(buffer, width=size, height=size)


def build_partner_request_excel(payload: dict[str, Any]) -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Requisicao"

    title_fill = PatternFill("solid", fgColor="0E2A47")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_fill = PatternFill("solid", fgColor="D4AF37")
    header_font = Font(bold=True, color="0E2A47")

    sheet.merge_cells("A1:D1")
    sheet["A1"] = payload["title"]
    sheet["A1"].fill = title_fill
    sheet["A1"].font = title_font
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet["A2"] = f"Chave PIX: {payload['pix_key']}"
    sheet["A3"] = payload["selected_period_label"] or "Todos os periodos"

    headers = ["Data", "Placa", "Servico", "Valor"]
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_index = 6
    for entry in payload["entries"]:
        sheet.cell(row=row_index, column=1, value=entry["display_date"])
        sheet.cell(row=row_index, column=2, value=entry["plate"])
        sheet.cell(row=row_index, column=3, value=entry["service_name"])
        sheet.cell(row=row_index, column=4, value=entry["amount"])
        row_index += 1

    sheet.cell(row=row_index + 1, column=3, value="Total")
    sheet.cell(row=row_index + 1, column=4, value=payload["total_value"])
    sheet.cell(row=row_index + 1, column=3).font = header_font
    sheet.cell(row=row_index + 1, column=4).font = header_font

    for column_letter, width in {"A": 16, "B": 22, "C": 34, "D": 16}.items():
        sheet.column_dimensions[column_letter].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_partner_request_pdf(payload: dict[str, Any]) -> io.BytesIO:
    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    pix_style = styles["Normal"].clone("RequestPixHeader")
    pix_style.textColor = colors.white
    pix_style.fontName = "Helvetica-Bold"
    pix_style.fontSize = 10
    pix_style.leading = 13
    pix_style.alignment = 1
    partner_style = styles["Title"].clone("RequestPartnerTitle")
    partner_style.textColor = colors.white
    partner_style.fontSize = 22
    partner_style.leading = 26
    partner_style.alignment = 1
    period_style = styles["Normal"].clone("RequestPeriodHeader")
    period_style.textColor = colors.HexColor("#F0D777")
    period_style.fontName = "Helvetica-Bold"
    period_style.fontSize = 10
    period_style.leading = 12
    period_style.alignment = 1
    story = []

    header = Table(
        [
            [
                Paragraph(f"Chave PIX: {payload['pix_key']}", pix_style),
                build_pix_qr_flowable(32 * mm),
            ],
            [
                Paragraph(f"<b>{payload['partner_name'].upper()}</b>", partner_style),
                "",
            ],
            [
                Paragraph(payload["selected_period_label"] or "Todos os periodos", period_style),
                "",
            ],
        ],
        colWidths=[140 * mm, 36 * mm],
    )
    header.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0E2A47")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#D4AF37")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("SPAN", (1, 0), (1, 2)),
                ("ALIGN", (0, 0), (0, 2), "CENTER"),
                ("ALIGN", (1, 0), (1, 2), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    story.extend([header, Spacer(1, 10)])

    table_data = [["Data", "Placa / Servico", "Valor"]]
    if payload["entries"]:
        for entry in payload["entries"]:
            plate_service = f"{entry['plate'] or '-'} / {entry['service_name']}"
            table_data.append([entry["display_date"], plate_service, f"R$ {entry['amount']:.2f}"])
    else:
        table_data.append(["-", "Sem requisicoes para este parceiro", "R$ 0.00"])
    table_data.append(["", "Total", f"R$ {payload['total_value']:.2f}"])

    table = Table(table_data, colWidths=[32 * mm, 104 * mm, 36 * mm], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D4AF37")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0E2A47")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D4AF37")),
                ("BACKGROUND", (0, 1), (-1, -2), colors.HexColor("#FFFDF7")),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#0E2A47")),
                ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (2, 1), (2, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]
        )
    )
    story.append(table)
    document.build(story)
    buffer.seek(0)
    return buffer


def init_db() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with get_db() as connection:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS app_meta (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS clients (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                client_name TEXT NOT NULL UNIQUE,
                price_transferencia REAL NOT NULL DEFAULT 0,
                price_caminhao_transferencia REAL NOT NULL DEFAULT 0,
                price_combo_transferencia REAL NOT NULL DEFAULT 0,
                price_cautelar REAL NOT NULL DEFAULT 0,
                price_pesquisa REAL NOT NULL DEFAULT 0,
                price_diversos REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS months (
                month_key TEXT PRIMARY KEY,
                year_number INTEGER NOT NULL,
                month_number INTEGER NOT NULL,
                month_label TEXT NOT NULL,
                month_title TEXT NOT NULL,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                month_key TEXT,
                reference_date TEXT,
                period_label TEXT NOT NULL,
                period_sort INTEGER NOT NULL DEFAULT 99,
                partner_name TEXT NOT NULL,
                transferencia_qty INTEGER NOT NULL DEFAULT 0,
                caminhao_transferencia_qty INTEGER NOT NULL DEFAULT 0,
                combo_transferencia_qty INTEGER NOT NULL DEFAULT 0,
                cautelar_qty INTEGER NOT NULL DEFAULT 0,
                pesquisa_qty INTEGER NOT NULL DEFAULT 0,
                unit_transferencia REAL NOT NULL DEFAULT 0,
                unit_caminhao_transferencia REAL NOT NULL DEFAULT 0,
                unit_combo_transferencia REAL NOT NULL DEFAULT 0,
                unit_cautelar REAL NOT NULL DEFAULT 0,
                unit_pesquisa REAL NOT NULL DEFAULT 0,
                total_value REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS cash_days (
                cash_date TEXT PRIMARY KEY,
                year_number INTEGER NOT NULL,
                month_number INTEGER NOT NULL,
                day_number INTEGER NOT NULL,
                month_title TEXT NOT NULL,
                finalized INTEGER NOT NULL DEFAULT 0,
                finalized_at TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS cash_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cash_date TEXT NOT NULL,
                customer_name TEXT NOT NULL,
                plate TEXT,
                service_name TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                payment_method TEXT NOT NULL,
                payment_group TEXT NOT NULL DEFAULT 'outras',
                flow_type TEXT NOT NULL DEFAULT 'entrada',
                request_payment_status TEXT NOT NULL DEFAULT 'em_aberto',
                synced_to_monthly INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (cash_date) REFERENCES cash_days (cash_date)
            );
            """
        )

        columns = {row["name"] for row in connection.execute("PRAGMA table_info(records)").fetchall()}
        if "month_key" not in columns:
            connection.execute("ALTER TABLE records ADD COLUMN month_key TEXT")
        if "combo_transferencia_qty" not in columns:
            connection.execute("ALTER TABLE records ADD COLUMN combo_transferencia_qty INTEGER NOT NULL DEFAULT 0")
        if "unit_combo_transferencia" not in columns:
            connection.execute("ALTER TABLE records ADD COLUMN unit_combo_transferencia REAL NOT NULL DEFAULT 0")
        if "caminhao_transferencia_qty" not in columns:
            connection.execute("ALTER TABLE records ADD COLUMN caminhao_transferencia_qty INTEGER NOT NULL DEFAULT 0")
        if "unit_caminhao_transferencia" not in columns:
            connection.execute("ALTER TABLE records ADD COLUMN unit_caminhao_transferencia REAL NOT NULL DEFAULT 0")
        client_columns = {row["name"] for row in connection.execute("PRAGMA table_info(clients)").fetchall()}
        if "price_transferencia" not in client_columns:
            connection.execute("ALTER TABLE clients ADD COLUMN price_transferencia REAL NOT NULL DEFAULT 0")
        if "price_caminhao_transferencia" not in client_columns:
            connection.execute("ALTER TABLE clients ADD COLUMN price_caminhao_transferencia REAL NOT NULL DEFAULT 0")
        if "price_combo_transferencia" not in client_columns:
            connection.execute("ALTER TABLE clients ADD COLUMN price_combo_transferencia REAL NOT NULL DEFAULT 0")
        if "price_cautelar" not in client_columns:
            connection.execute("ALTER TABLE clients ADD COLUMN price_cautelar REAL NOT NULL DEFAULT 0")
        if "price_pesquisa" not in client_columns:
            connection.execute("ALTER TABLE clients ADD COLUMN price_pesquisa REAL NOT NULL DEFAULT 0")
        if "price_diversos" not in client_columns:
            connection.execute("ALTER TABLE clients ADD COLUMN price_diversos REAL NOT NULL DEFAULT 0")
        cash_columns = {row["name"] for row in connection.execute("PRAGMA table_info(cash_entries)").fetchall()}
        if "request_payment_status" not in cash_columns:
            connection.execute("ALTER TABLE cash_entries ADD COLUMN request_payment_status TEXT NOT NULL DEFAULT 'em_aberto'")
        if "synced_to_monthly" not in cash_columns:
            connection.execute("ALTER TABLE cash_entries ADD COLUMN synced_to_monthly INTEGER NOT NULL DEFAULT 0")
        connection.execute(
            """
            UPDATE cash_entries
            SET request_payment_status = 'em_aberto'
            WHERE request_payment_status IS NULL OR TRIM(request_payment_status) = ''
            """
        )

        connection.executescript(
            """
            CREATE INDEX IF NOT EXISTS idx_records_partner ON records (partner_name);
            CREATE INDEX IF NOT EXISTS idx_records_month ON records (month_key);
            CREATE INDEX IF NOT EXISTS idx_months_sort ON months (year_number, month_number);
            CREATE INDEX IF NOT EXISTS idx_clients_name ON clients (client_name);
            CREATE INDEX IF NOT EXISTS idx_cash_entries_date ON cash_entries (cash_date);
            CREATE INDEX IF NOT EXISTS idx_cash_entries_customer ON cash_entries (customer_name);
            CREATE INDEX IF NOT EXISTS idx_cash_days_sort ON cash_days (year_number, month_number, day_number);
            DROP TABLE IF EXISTS private_clients;
            """
        )

        record_count = connection.execute("SELECT COUNT(*) FROM records").fetchone()[0]
        if record_count == 0 and SEED_PATH.exists():
            seed_records = json.loads(SEED_PATH.read_text(encoding="utf-8-sig"))
            seed_month_map = infer_seed_month_map(seed_records)
            for month_key in sorted(set(seed_month_map.values())):
                if is_month_allowed(month_key):
                    year_number, month_number = parse_month_key(month_key)
                    ensure_month(connection, year_number, month_number)

            connection.executemany(
                """
                INSERT INTO records (
                    month_key,
                    reference_date,
                    period_label,
                    period_sort,
                    partner_name,
                    transferencia_qty,
                    caminhao_transferencia_qty,
                    combo_transferencia_qty,
                    cautelar_qty,
                    pesquisa_qty,
                    unit_transferencia,
                    unit_caminhao_transferencia,
                    unit_combo_transferencia,
                    unit_cautelar,
                    unit_pesquisa,
                    total_value
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [
                    (
                        seed_month_map.get(normalize_text(item["period_label"]), MIN_MONTH_KEY),
                        None,
                        normalize_text(item["period_label"]),
                        0,
                        item["partner_name"],
                        item["transferencia_qty"],
                        0,
                        0,
                        item["cautelar_qty"],
                        item["pesquisa_qty"],
                        item["unit_transferencia"],
                        0,
                        0,
                        item["unit_cautelar"],
                        item["unit_pesquisa"],
                        item["total_value"],
                    )
                    for item in seed_records
                    if is_month_allowed(seed_month_map.get(normalize_text(item["period_label"]), MIN_MONTH_KEY))
                ],
            )

        missing_month_rows = connection.execute(
            """
            SELECT period_label, MIN(id) AS first_id
            FROM records
            WHERE month_key IS NULL OR month_key = ''
            GROUP BY period_label
            ORDER BY first_id
            """
        ).fetchall()
        if missing_month_rows:
            inferred_map = infer_seed_month_map([{"period_label": row["period_label"]} for row in missing_month_rows])
            for label, month_key in inferred_map.items():
                if is_month_allowed(month_key):
                    year_number, month_number = parse_month_key(month_key)
                    ensure_month(connection, year_number, month_number)
                    connection.execute(
                        """
                        UPDATE records
                        SET month_key = ?
                        WHERE (month_key IS NULL OR month_key = '')
                          AND period_label = ?
                        """,
                        (month_key, label),
                    )

        existing_bound_months = connection.execute(
            "SELECT DISTINCT month_key FROM records WHERE month_key IS NOT NULL AND month_key <> ''"
        ).fetchall()
        for row in existing_bound_months:
            if is_month_allowed(row["month_key"]):
                year_number, month_number = parse_month_key(row["month_key"])
                ensure_month(connection, year_number, month_number)

        connection.execute("DELETE FROM records WHERE month_key IS NULL OR month_key = ''")
        connection.execute("DELETE FROM records WHERE month_key < ? OR month_key > ?", (MIN_MONTH_KEY, MAX_MONTH_KEY))
        cleanup_done = connection.execute(
            "SELECT value FROM app_meta WHERE key = 'cleanup_records_2026_04_2026_11'"
        ).fetchone()
        if cleanup_done is None:
            connection.execute(
                "DELETE FROM records WHERE month_key >= ? AND month_key <= ?",
                ("2026-04", "2026-11"),
            )
            connection.execute(
                "INSERT INTO app_meta (key, value) VALUES ('cleanup_records_2026_04_2026_11', 'done')"
            )
        ensure_allowed_months(connection)


@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""
    next_url = request.args.get("next") or url_for("index")
    if request.method == "POST":
        password = request.form.get("password", "")
        next_url = request.form.get("next") or url_for("index")
        if not APP_PASSWORD:
            error = "Senha do sistema nao configurada. Defina APP_PASSWORD no ambiente."
            return render_template("login.html", error=error, next_url=next_url), 500
        if hmac.compare_digest(password, APP_PASSWORD):
            session.clear()
            session["authenticated"] = True
            safe_next_url = next_url if next_url.startswith("/") and not next_url.startswith("//") else url_for("index")
            return redirect(safe_next_url)
        error = "Senha incorreta. Tente novamente."
    return render_template("login.html", error=error, next_url=next_url)


@app.post("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
def index() -> str:
    return render_template("index.html")


@app.route("/service-order")
def service_order_page() -> str:
    return render_template("service_order.html")


@app.route("/cash-flow")
def cash_flow_page() -> str:
    return render_template("cash_flow.html")


@app.route("/comparison")
def comparison_page() -> str:
    return render_template("comparison.html")


@app.route("/client-comparison")
def client_comparison_page() -> str:
    return render_template("client_comparison.html")


@app.route("/partner-requests")
def partner_requests_page() -> str:
    return render_template("partner_requests.html")


@app.get("/healthz")
def healthcheck():
    return jsonify({"status": "ok"})


@app.get("/api/records")
def list_records():
    search = (request.args.get("search") or "").strip()
    month_key = clamp_month_key((request.args.get("month_key") or get_default_month_key()).strip())
    sort_by = ALLOWED_SORT_COLUMNS.get(request.args.get("sort_by"), "partner_name")
    sort_order = "DESC" if (request.args.get("sort_order") or "").lower() == "desc" else "ASC"

    try:
        year_number, month_number = parse_month_key(month_key)
        if not is_month_allowed(month_key):
            raise ValueError("O mes precisa estar entre Abril de 2026 e Dezembro de 2050.")
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    with get_db() as connection:
        ensure_month(connection, year_number, month_number)
        ensure_allowed_months(connection)
        report = get_month_report(connection, month_key, search, sort_by, sort_order)
        months = list_months(connection)
        active_month = next((item for item in months if item["month_key"] == month_key), None)

    return jsonify(
        {
            "records": [serialize_main_record(row) for row in report["records"]],
            "months": months,
            "summary": report["summary"],
            "active_month": active_month,
            "meta": {"total_records": len(report["records"])},
        }
    )


@app.get("/api/comparison")
def comparison_data():
    with get_db() as connection:
        ensure_allowed_months(connection)
        months = get_comparison_data(connection)
        clients = get_all_clients(connection)
    return jsonify({"months": months, "clients": clients})


@app.get("/api/comparison/<string:month_key>")
def comparison_month_detail(month_key: str):
    month_key = clamp_month_key(month_key)
    try:
        year_number, month_number = parse_month_key(month_key)
        if not is_month_allowed(month_key):
            raise ValueError("Mes fora do intervalo permitido.")
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    with get_db() as connection:
        ensure_month(connection, year_number, month_number)
        payload = get_month_chart_data(connection, month_key)
    return jsonify(payload)


@app.get("/api/client-history/<path:partner_name>")
def client_history(partner_name: str):
    if not partner_name.strip():
        return jsonify({"error": "Cliente invalido."}), 400
    with get_db() as connection:
        payload = get_client_history(connection, partner_name)
    return jsonify(payload)


@app.get("/api/month-compare")
def month_compare():
    first = clamp_month_key((request.args.get("first") or "").strip())
    second = clamp_month_key((request.args.get("second") or "").strip())
    if not first or not second:
        return jsonify({"error": "Selecione dois meses para comparar."}), 400

    try:
        parse_month_key(first)
        parse_month_key(second)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    with get_db() as connection:
        ensure_allowed_months(connection)
        payload = compare_two_months(connection, first, second)
    return jsonify(payload)


@app.get("/api/client-comparison")
def client_comparison():
    month_key = (request.args.get("month_key") or "").strip()
    if month_key:
        month_key = clamp_month_key(month_key)
        try:
            parse_month_key(month_key)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    else:
        month_key = None

    with get_db() as connection:
        ensure_allowed_months(connection)
        payload = get_client_comparison_payload(connection, month_key)
        months = get_comparison_data(connection)
    return jsonify({"months": months, **payload})


@app.get("/api/clients")
def list_clients():
    with get_db() as connection:
        sync_clients_catalog(connection)
        cleanup_clients_catalog(connection)
        clients = get_all_clients(connection)
        items = get_client_catalog(connection)
    return jsonify({"clients": clients, "items": items, "total_count": len(items)})


@app.post("/api/clients")
def create_client():
    payload = request.get_json(silent=True) or {}
    client_name = payload.get("client_name", "")
    try:
        with get_db() as connection:
            saved_client = register_client(connection, client_name, payload)
            cleanup_clients_catalog(connection)
            clients = get_all_clients(connection)
            items = get_client_catalog(connection)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify(
        {"message": "Cliente cadastrado com sucesso.", "client": saved_client, "clients": clients, "items": items, "total_count": len(items)},
    ), 201


@app.put("/api/clients/<int:client_id>")
def edit_client(client_id: int):
    payload = request.get_json(silent=True) or {}
    client_name = payload.get("client_name", "")
    try:
        with get_db() as connection:
            saved_client = update_client(connection, client_id, client_name, payload)
            cleanup_clients_catalog(connection)
            clients = get_all_clients(connection)
            items = get_client_catalog(connection)
    except ValueError as exc:
        status_code = 404 if "nao encontrado" in str(exc).lower() else 400
        return jsonify({"error": str(exc)}), status_code
    return jsonify(
        {"message": "Cliente atualizado com sucesso.", "client": saved_client, "clients": clients, "items": items, "total_count": len(items)}
    )


@app.delete("/api/clients/<int:client_id>")
def delete_client(client_id: int):
    with get_db() as connection:
        existing = connection.execute("SELECT id FROM clients WHERE id = ?", (client_id,)).fetchone()
        if existing is None:
            return jsonify({"error": "Cliente nao encontrado."}), 404
        connection.execute("DELETE FROM clients WHERE id = ?", (client_id,))
        cleanup_clients_catalog(connection)
        clients = get_all_clients(connection)
        items = get_client_catalog(connection)
    return jsonify({"message": "Cliente excluido com sucesso.", "clients": clients, "items": items, "total_count": len(items)})


@app.post("/api/clients/cleanup")
def cleanup_clients():
    with get_db() as connection:
        sync_clients_catalog(connection)
        stats = cleanup_clients_catalog(connection)
        clients = get_all_clients(connection)
        items = get_client_catalog(connection)
    return jsonify(
        {
            "message": "Clientes invalidados/ocultos foram limpos com sucesso.",
            "stats": stats,
            "clients": clients,
            "items": items,
            "total_count": len(items),
        }
    )


@app.get("/api/partner-requests")
def partner_requests():
    with get_db() as connection:
        payload = get_partner_requests_payload(connection)
    return jsonify(payload)


@app.get("/api/partner-requests/detail")
def partner_request_detail():
    partner_name = (request.args.get("partner") or "").strip()
    if not partner_name:
        return jsonify({"error": "Parceiro obrigatorio."}), 400
    try:
        year_number, month_number = parse_request_period(request.args.get("year"), request.args.get("month"))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    with get_db() as connection:
        payload = get_partner_request_detail(connection, partner_name, year_number=year_number, month_number=month_number)
    return jsonify(payload)


@app.put("/api/partner-requests/status/<int:entry_id>")
def update_partner_request_status(entry_id: int):
    status = normalize_request_payment_status((request.get_json(silent=True) or {}).get("status"))
    with get_db() as connection:
        row = connection.execute(
            """
            SELECT id
            FROM cash_entries
            WHERE id = ? AND UPPER(TRIM(payment_method)) = 'REC'
            """,
            (entry_id,),
        ).fetchone()
        if row is None:
            return jsonify({"error": "Requisicao nao encontrada."}), 404
        connection.execute(
            """
            UPDATE cash_entries
            SET request_payment_status = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (status, entry_id),
        )
    return jsonify({"message": "Status atualizado com sucesso.", "status": status, "status_label": request_payment_status_label(status)})


@app.get("/api/partner-requests/pix-qr.png")
def partner_request_pix_qr():
    buffer = build_pix_qr_png()
    return send_file(buffer, mimetype="image/png", as_attachment=False, download_name="pix-qrcode.png")


@app.get("/api/partner-requests/pix-code")
def partner_request_pix_code():
    return jsonify({"pix_key": PIX_KEY, "pix_copy_paste": build_pix_payload()})


@app.get("/api/partner-requests/export.xlsx")
def export_partner_request_xlsx():
    partner_name = (request.args.get("partner") or "").strip()
    if not partner_name:
        return jsonify({"error": "Parceiro obrigatorio."}), 400
    try:
        year_number, month_number = parse_request_period(request.args.get("year"), request.args.get("month"))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    with get_db() as connection:
        payload = get_partner_request_detail(connection, partner_name, year_number=year_number, month_number=month_number)
    buffer = build_partner_request_excel(payload)
    safe_name = safe_filename_part(partner_name)
    suffix = f"-{year_number:04d}-{month_number:02d}" if year_number and month_number else ""
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"requisicao-{safe_name}{suffix}.xlsx",
    )


@app.get("/api/partner-requests/export.pdf")
def export_partner_request_pdf():
    partner_name = (request.args.get("partner") or "").strip()
    if not partner_name:
        return jsonify({"error": "Parceiro obrigatorio."}), 400
    try:
        year_number, month_number = parse_request_period(request.args.get("year"), request.args.get("month"))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    with get_db() as connection:
        payload = get_partner_request_detail(
            connection,
            partner_name,
            year_number=year_number,
            month_number=month_number,
            only_open=True,
        )
    buffer = build_partner_request_pdf(payload)
    safe_name = safe_filename_part(partner_name)
    suffix = f"-{year_number:04d}-{month_number:02d}" if year_number and month_number else ""
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"requisicao-{safe_name}{suffix}.pdf",
    )


@app.get("/api/cash-flow/tree")
def cash_flow_tree():
    with get_db() as connection:
        return jsonify({"tree": get_cash_tree(connection)})


@app.get("/api/cash-flow/day/<string:cash_date>")
def cash_flow_day(cash_date: str):
    try:
        parse_cash_date(cash_date)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    with get_db() as connection:
        payload = get_cash_day_payload(connection, cash_date)
    return jsonify(payload)


@app.post("/api/cash-flow/day/<string:cash_date>/entries")
def create_cash_entry(cash_date: str):
    try:
        parse_cash_date(cash_date)
        entry = validate_cash_entry_payload(request.get_json(silent=True) or {})
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    with get_db() as connection:
        ensure_cash_day(connection, cash_date)
        upsert_client_from_system(connection, entry["customer_name"])
        cursor = connection.execute(
            """
            INSERT INTO cash_entries (
                cash_date,
                customer_name,
                plate,
                service_name,
                amount,
                payment_method,
                payment_group,
                flow_type,
                synced_to_monthly,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, CURRENT_TIMESTAMP)
            """,
            (
                cash_date,
                entry["customer_name"],
                entry["plate"],
                entry["service_name"],
                entry["amount"],
                entry["payment_method"],
                entry["payment_group"],
                entry["flow_type"],
            ),
        )
        row = connection.execute("SELECT * FROM cash_entries WHERE id = ?", (cursor.lastrowid,)).fetchone()
        sync_cash_entry_to_monthly(connection, row, 1)
        connection.execute("UPDATE cash_entries SET synced_to_monthly = 1 WHERE id = ?", (row["id"],))
        connection.execute("UPDATE cash_days SET updated_at = CURRENT_TIMESTAMP WHERE cash_date = ?", (cash_date,))
        row = connection.execute("SELECT * FROM cash_entries WHERE id = ?", (cursor.lastrowid,)).fetchone()
    return jsonify(serialize_cash_entry(row)), 201


@app.put("/api/cash-flow/entries/<int:entry_id>")
def update_cash_entry(entry_id: int):
    try:
        entry = validate_cash_entry_payload(request.get_json(silent=True) or {})
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    with get_db() as connection:
        existing = connection.execute("SELECT * FROM cash_entries WHERE id = ?", (entry_id,)).fetchone()
        if existing is None:
            return jsonify({"error": "Lancamento nao encontrado."}), 404
        if existing["synced_to_monthly"]:
            sync_cash_entry_to_monthly(connection, existing, -1)
        upsert_client_from_system(connection, entry["customer_name"])
        connection.execute(
            """
            UPDATE cash_entries
            SET
                customer_name = ?,
                plate = ?,
                service_name = ?,
                amount = ?,
                payment_method = ?,
                payment_group = ?,
                flow_type = ?,
                synced_to_monthly = 0,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                entry["customer_name"],
                entry["plate"],
                entry["service_name"],
                entry["amount"],
                entry["payment_method"],
                entry["payment_group"],
                entry["flow_type"],
                entry_id,
            ),
        )
        row = connection.execute("SELECT * FROM cash_entries WHERE id = ?", (entry_id,)).fetchone()
        sync_cash_entry_to_monthly(connection, row, 1)
        connection.execute("UPDATE cash_entries SET synced_to_monthly = 1 WHERE id = ?", (entry_id,))
        connection.execute("UPDATE cash_days SET updated_at = CURRENT_TIMESTAMP WHERE cash_date = ?", (existing["cash_date"],))
        row = connection.execute("SELECT * FROM cash_entries WHERE id = ?", (entry_id,)).fetchone()
    return jsonify(serialize_cash_entry(row))


@app.delete("/api/cash-flow/entries/<int:entry_id>")
def delete_cash_entry(entry_id: int):
    with get_db() as connection:
        existing = connection.execute("SELECT * FROM cash_entries WHERE id = ?", (entry_id,)).fetchone()
        if existing is None:
            return jsonify({"error": "Lancamento nao encontrado."}), 404
        if existing["synced_to_monthly"]:
            sync_cash_entry_to_monthly(connection, existing, -1)
        connection.execute("DELETE FROM cash_entries WHERE id = ?", (entry_id,))
        connection.execute("UPDATE cash_days SET updated_at = CURRENT_TIMESTAMP WHERE cash_date = ?", (existing["cash_date"],))
    return jsonify({"message": "Lancamento excluido com sucesso."})


@app.delete("/api/cash-flow/day/<string:cash_date>")
def delete_cash_day(cash_date: str):
    try:
        parse_cash_date(cash_date)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    with get_db() as connection:
        day = connection.execute("SELECT cash_date FROM cash_days WHERE cash_date = ?", (cash_date,)).fetchone()
        if day is None:
            return jsonify({"error": "Caixa nao encontrado."}), 404
        entries = connection.execute("SELECT * FROM cash_entries WHERE cash_date = ?", (cash_date,)).fetchall()
        for entry in entries:
            if entry["synced_to_monthly"]:
                sync_cash_entry_to_monthly(connection, entry, -1)
        connection.execute("DELETE FROM cash_entries WHERE cash_date = ?", (cash_date,))
        connection.execute("DELETE FROM cash_days WHERE cash_date = ?", (cash_date,))
    return jsonify({"message": "Caixa excluido com sucesso."})


@app.post("/api/cash-flow/day/<string:cash_date>/finalize")
def finalize_cash_day(cash_date: str):
    try:
        parse_cash_date(cash_date)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    with get_db() as connection:
        ensure_cash_day(connection, cash_date)
        unsynced_rows = connection.execute(
            "SELECT * FROM cash_entries WHERE cash_date = ? AND synced_to_monthly = 0",
            (cash_date,),
        ).fetchall()
        for row in unsynced_rows:
            sync_cash_entry_to_monthly(connection, row, 1)
            connection.execute("UPDATE cash_entries SET synced_to_monthly = 1 WHERE id = ?", (row["id"],))
        connection.execute(
            """
            UPDATE cash_days
            SET finalized = 1, finalized_at = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP
            WHERE cash_date = ?
            """,
            (cash_date,),
        )
        payload = get_cash_day_payload(connection, cash_date)
    return jsonify(payload)


@app.post("/api/cash-flow/day/<string:cash_date>/reopen")
def reopen_cash_day(cash_date: str):
    try:
        parse_cash_date(cash_date)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    with get_db() as connection:
        ensure_cash_day(connection, cash_date)
        connection.execute(
            """
            UPDATE cash_days
            SET finalized = 0, finalized_at = NULL, updated_at = CURRENT_TIMESTAMP
            WHERE cash_date = ?
            """,
            (cash_date,),
        )
        payload = get_cash_day_payload(connection, cash_date)
    return jsonify(payload)


@app.get("/api/cash-flow/day/<string:cash_date>.pdf")
def export_cash_day_pdf(cash_date: str):
    try:
        parse_cash_date(cash_date)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    with get_db() as connection:
        payload = get_cash_day_payload(connection, cash_date)
    buffer = build_cash_pdf_report(payload)
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"fluxo-caixa-{cash_date}.pdf",
    )


@app.get("/api/cash-flow/month/<string:month_key>.pdf")
def export_cash_month_pdf(month_key: str):
    try:
        parse_month_key(month_key)
        if not is_month_allowed(month_key):
            raise ValueError("Mes fora do intervalo permitido.")
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    with get_db() as connection:
        payload = get_cash_month_payload(connection, month_key)
    buffer = build_cash_month_pdf_report(payload)
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"fluxo-caixa-mensal-{month_key}.pdf",
    )


@app.get("/api/export/<string:month_key>.xlsx")
def export_month_xlsx(month_key: str):
    month_key = clamp_month_key(month_key)
    try:
        year_number, month_number = parse_month_key(month_key)
        if not is_month_allowed(month_key):
            raise ValueError("Mes fora do intervalo permitido.")
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    with get_db() as connection:
        ensure_month(connection, year_number, month_number)
        report = get_month_report(connection, month_key)

    buffer = build_excel_report(report)
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"relatorio-mensal-{month_key}.xlsx",
    )


@app.get("/api/export/<string:month_key>.pdf")
def export_month_pdf(month_key: str):
    month_key = clamp_month_key(month_key)
    try:
        year_number, month_number = parse_month_key(month_key)
        if not is_month_allowed(month_key):
            raise ValueError("Mes fora do intervalo permitido.")
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    with get_db() as connection:
        ensure_month(connection, year_number, month_number)
        report = get_month_report(connection, month_key)

    buffer = build_pdf_report(report)
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"relatorio-mensal-{month_key}.pdf",
    )


@app.post("/api/records")
def create_record():
    payload = request.get_json(silent=True) or {}
    try:
        record = validate_main_payload(payload)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    with get_db() as connection:
        ensure_month(connection, record["year_number"], record["month_number"])
        upsert_client_from_system(connection, record["partner_name"])
        cursor = connection.execute(
            """
            INSERT INTO records (
                month_key,
                reference_date,
                period_label,
                period_sort,
                partner_name,
                transferencia_qty,
                caminhao_transferencia_qty,
                combo_transferencia_qty,
                cautelar_qty,
                pesquisa_qty,
                unit_transferencia,
                unit_caminhao_transferencia,
                unit_combo_transferencia,
                unit_cautelar,
                unit_pesquisa,
                total_value,
                updated_at
            ) VALUES (?, NULL, ?, 0, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            """,
            (
                record["month_key"],
                record["period_label"],
                record["partner_name"],
                record["transferencia_qty"],
                record["caminhao_transferencia_qty"],
                record["combo_transferencia_qty"],
                record["cautelar_qty"],
                record["pesquisa_qty"],
                record["unit_transferencia"],
                record["unit_caminhao_transferencia"],
                record["unit_combo_transferencia"],
                record["unit_cautelar"],
                record["unit_pesquisa"],
                record["total_value"],
            ),
        )
        row = connection.execute("SELECT * FROM records WHERE id = ?", (cursor.lastrowid,)).fetchone()

    return jsonify(serialize_main_record(row)), 201


@app.put("/api/records/<int:record_id>")
def update_record(record_id: int):
    payload = request.get_json(silent=True) or {}
    try:
        record = validate_main_payload(payload)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    with get_db() as connection:
        existing = connection.execute("SELECT id FROM records WHERE id = ?", (record_id,)).fetchone()
        if existing is None:
            return jsonify({"error": "Registro nao encontrado."}), 404

        ensure_month(connection, record["year_number"], record["month_number"])
        upsert_client_from_system(connection, record["partner_name"])
        connection.execute(
            """
            UPDATE records
            SET
                month_key = ?,
                reference_date = NULL,
                period_label = ?,
                partner_name = ?,
                transferencia_qty = ?,
                caminhao_transferencia_qty = ?,
                combo_transferencia_qty = ?,
                cautelar_qty = ?,
                pesquisa_qty = ?,
                unit_transferencia = ?,
                unit_caminhao_transferencia = ?,
                unit_combo_transferencia = ?,
                unit_cautelar = ?,
                unit_pesquisa = ?,
                total_value = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                record["month_key"],
                record["period_label"],
                record["partner_name"],
                record["transferencia_qty"],
                record["caminhao_transferencia_qty"],
                record["combo_transferencia_qty"],
                record["cautelar_qty"],
                record["pesquisa_qty"],
                record["unit_transferencia"],
                record["unit_caminhao_transferencia"],
                record["unit_combo_transferencia"],
                record["unit_cautelar"],
                record["unit_pesquisa"],
                record["total_value"],
                record_id,
            ),
        )
        row = connection.execute("SELECT * FROM records WHERE id = ?", (record_id,)).fetchone()

    return jsonify(serialize_main_record(row))


@app.delete("/api/records/<int:record_id>")
def delete_record(record_id: int):
    with get_db() as connection:
        cursor = connection.execute("DELETE FROM records WHERE id = ?", (record_id,))
        if cursor.rowcount == 0:
            return jsonify({"error": "Registro nao encontrado."}), 404
    return jsonify({"message": "Registro excluido com sucesso."})


init_db()


if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    debug = os.getenv("FLASK_DEBUG", "").lower() in {"1", "true", "yes"}
    app.run(host="0.0.0.0", port=port, debug=debug)
